import io `DD/MM/AAAA`** — tanto a vigência quanto a competência fim

Aqu
import re
import time
import unicodedata
import requests
importi está o código completo final com todas as correções aplicadas sobre o `p pandas as pd
import streamlit as st
from datetime import date

TR_ORANGE      = "#FFasted-text` (versão base que8000"
TR_ORANGE_DARK = "#CC6600"
TR_BG          = "#1 você enviou):

```python
import io
import re
import time
import unicA1A1A"
TR_CARD        = "#242424"
TR_CARD2       = "#2E2E2E"
TRodedata
import requests
import pandas as pd
import streamlit as st
from datetime import date

TR_BORDER      = "#3A3A3A"
TR_TEXT        = "#F0_ORANGE      = "#FF8000"
TR_ORANGEF0F0"
TR_TEXT_MUTED  = "#999999"
TR_SUCCESS     = "#2ECC71"
TR_ERROR_DARK = "#CC6600"
TR_BG          = "#1A1A1A"
TR_CARD        = "#242424"
TR_CARD       = "#E74C3C"
TR_WARNING     = "#F39C12"

CSS = f"""
<style>
  2       = "#2E2E2E"
TR_BORDER      = "#3A3A3A"
TR_TEXThtml, body, [data-testid="stAppViewContainer"] {{
      background-color: {TR_B        = "#F0F0F0"
TR_TEXT_MUTED  = "#999999"
TR_SUCCESS     = "#2ECC71"
TR_G} !important;
      color: {TR_TEXT} !important;
      font-family: 'Segoe UI', Arial, sans-serif !important;
  }}
  section[data-testid="stSidebar"] {{
      background-color: #111111 !important;
      borderERROR       = "#E74C3C"
TR_WARNING     = "#F39C12"

CSS = f"""
<style>
  -right: 2px solid {TR_ORANGE} !important;
      display: block !important;
      visibility: visible !important;
      min-width: 21rem !important;
      transform: translateX(0) !important;
      left: 0 !important;
  }}
  section[data-testid="stSidebar"] > div:first-child {{ background-color: #111111 !important; }}
  section[data-testid="stSidebar"] * {{ color: {TR_TEXT} !important; }}
  [data-testid="collapsedControl"] {{
      display: flex !important; visibility: visible !important;
      opacity: 1 !important; color: {TR_ORANGE} !important;
  }}
  .tr-header {{
      background: linear-gradient(135deg,#111 0%,#1f1f1f 60%,#2a1500 100%);
      border-bottom: 3px solid {TR_ORANGE}; padding: 18px 28px 14html, body, [data-testid="stAppViewContainer"] {{
      background-color: {TR_BG} !important;
      color: {TR_TEXT} !important;
      font-family: 'Segoe UI', Arial, sans-serif !important;
  }}
  section[data-testid="stSidebar"] {{
      background-color: #111111 !important;
      border-right: 2px solid {TR_ORANGE} !important;
      display: block !important;
      visibility: visible !important;
      min-width: 21rem !important;
      transform: translateX(0) !important;
      left: 0 !important;
  }}
  section[data-testid="stSidebar"] > div:first-child {{ background-color: #111111 !important; }}
  section[data-testid="stSidebar"] * {{ color: {TR_TEXT} !important; }}
  [data-testid="collapsedControl"] {{
      display: flex !important; visibility: visible !important;
      opacity: 1 !important; color: {TR_ORANGE} !important;
  }}
  .tr-header {{
      background: linear-gradient(135deg,#111 0%,#1fpx;
      border-radius: 10px; margin-bottom: 24px; display: flex; align-items: center; gap: 181f1f 60%,#2a1500 100%);
      borderpx;
  }}
  .tr-logo-box {{
      background:{TR_ORANGE}; border-bottom: 3px solid {TR_ORANGE}; padding: 18px 28px 14px;
      border-radius:8px; width:52px; height:52px;
      display:flex; align-items:center; justify-content:center;
      font-radius: 10px; margin-bottom: 24px; display: flex; align-items: center; gap: 18px;
  }}
  -size:26px; font-weight:900; color:#fff; flex-shrink:0;
  }}
  .tr-title   .tr-logo-box {{
      background:{TR_ORANGE}; border-radius:{{ font-size:22px; font-weight:700; color:{TR_TEXT}; margin:0; line-height:1.2; }}
  .tr-subtitle {{ font-size:128px; width:52px; height:52px;
      display:flex; align-items:center; justify-content:center;
      font-size:26px; font-weight:900px; color:{TR_TEXT_MUTED}; margin:2px 0 0; letter-spacing:.5px; }}
  .tr-badge; color:#fff; flex-shrink:0;
  }}
  .tr-title   {{ font-size:22px; font-weight:700; color:{TR_TEXT {{
      margin-left:auto; background:{TR_ORANGE}; color:#fff; font-size:10px; font-weight:700;
      }; margin:0; line-height:1.2; }}
  .tr-subtitle {{ font-size:12px; color:{TR_TEXT_MUTED}; margin:2padding:4px 10px; border-radius:20px; letter-spacing:1px; text-transform:uppercase;
  }}
  .tr-cardpx 0 0; letter-spacing:.5px; }}
  .tr-badge {{
      margin-left: {{
      background:{TR_CARD}; border:1px solid {TR_BORDER};
      border-radius:10px; padding:18px 20px; margin-bottom:16auto; background:{TR_ORANGE}; color:#fff; font-size:10px; font-weight:700;
      padding:4px 10px; border-radius:20px;
  }}
  .tr-card-title {{
      font-size:13px; font-weight:700; color:{TR_ORANGE}; textpx; letter-spacing:1px; text-transform:uppercase;
  }}
  .tr-card {{
      background:{TR_CARD}; border:-transform:uppercase;
      letter-spacing:1px; margin-bottom:12px; border-bottom:1px solid {TR_BORDER}; padding-bottom:8px;
  }}
  .1px solid {TR_BORDER};
      border-radius:10px; padding:18px 20px; margin-bottom:16px;
  }}
  .tr-cardtr-metrics {{ display:flex; gap:12px; flex-wrap:wrap; margin-bottom:20px; }}
  .tr-title {{
      font-size:13px; font-weight:700; color:{TR_ORANGE}; text-transform:uppercase;
      letter-spacing:1px; margin-bottom:12-metric {{
      background:{TR_CARD}; border:1px solid {TR_BORDER}; border-radius:10px;
      padding:14px 20px; flex:px; border-bottom:1px solid {TR_BORDER}; padding-bottom:8px;
  }}
  .tr-metrics {{ display:1; min-width:120px; text-align:center;
  }}
  .tr-metric-value {{ font-size:28px; font-weight:800; color:{TR_flex; gap:12px; flex-wrap:wrap; margin-bottom:20px; }}
  .tr-metric {{
      background:{TR_CARD}; border:ORANGE}; line-height:1; }}
  .tr-metric-label {{ font-size:11px; color:{TR_TEXT_MUTED}; margin-top:4px; text1px solid {TR_BORDER}; border-radius:10px;
      padding:14px 20px; flex:1; min-width:120px; text-align:center;-transform:uppercase; letter-spacing:.5px; }}
  .tr-metric.success .tr-metric-value {{ color:{TR_SUCCESS}; }}
  .tr-
  }}
  .tr-metric-value {{ font-size:28px; font-weight:800; color:{TR_ORANGE}; line-height:1; }}
  .tr-metricmetric.error   .tr-metric-value {{ color:{TR_ERROR};   }}
  .tr-metric.warning .tr-metric-value {{ color:{TR_WARNING}; }}
  [data-testid="st-label {{ font-size:11px; color:{TR_TEXT_MUTED}; margin-top:4px; text-transform:uppercase; letter-spacing:.5px; }}
  .DataFrame"] {{ border:1px solid {TR_BORDER} !important; border-radius:8px !tr-metric.success .tr-metric-value {{ color:{TR_SUCCESS}; }}
  .tr-metric.error   .tr-metric-value {{important; }}
  .stButton > button {{
      background:{TR_ORANGE} !important; color:#fff ! color:{TR_ERROR};   }}
  .tr-metric.warning .tr-metric-value {{ color:{TR_WARNING}; }}
  [data-testid="stDataFrameimportant; border:none !important;
      border-radius:6px !important; font-weight:700 !important; letter-spacing:.5px !"] {{ border:1px solid {TR_BORDER} !important; border-radius:8px !important; }}
  .important;
  }}
  .stButton > button:hover {{ background:{TR_ORANGE_DARK} !important; }}
  .stDownloadButton > button {{
      background:#stButton > button {{
      background:{TR_ORANGE} !important; color:#fff !important; border:none !important1a3a1a !important; color:{TR_SUCCESS} !important;
      border:1px solid {TR_SUCCESS} !important; border-radius:6px !important; font-weight:700 !important;
  }}
  .stDownloadButton > button:hover {{ background:{TR_SUCCESS} !important;;
      border-radius:6px !important; font-weight:700 !important; letter-spacing:.5px !important;
  }}
  .stButton > button:hover {{ background:{TR_ORANGE_DARK} !important; }}
  .stDownloadButton > button {{
      background:# color:#fff !important; }}
  .stTextInput input, .stNumberInput input {{
      background:{TR_CARD21a3a1a !important; color:{TR_SUCCESS} !important;
      border:1px solid {TR_SUCCESS} !important; color:{TR_TEXT} !important;
      border:1px solid {TR_BORDER} !important; border-radius:6px !important;
  } !important; border-radius:6px !important; font-weight:700 !important;
  }}
  .stDownloadButton > button:hover {{ background:{TR_SUCCESS} !important; color}}
  .stTextArea textarea {{
      background:{TR_CARD2} !important; color:{TR_TEXT} !important;
      border:1px solid {TR_BORDER} !important; border-radius:6px !important;:#fff !important; }}
  .stTextInput input, .stNumberInput input {{
      background:{TR_CARD2
      font-family:'Courier New',monospace !important; font-size:13px !important;} !important; color:{TR_TEXT} !important;
      border:1px solid {TR_BORDER} !important; border-radius:6px !important;
  }}
  .stTextArea
  }}
  .stTextArea textarea:focus {{ border-color:{TR_ORANGE} !important; box-shadow:0 0 0 2px rgba textarea {{
      background:{TR_CARD2} !important; color:{TR_TEXT} !important;
      border:1px solid {TR_BORDER} !important; border-radius:6px !important;
      (255,128,0,.2) !important; }}
  .stMultiSelect [data-baseweb="selectfont-family:'Courier New',monospace !important; font-size:13px !important;
  }}
  .st"], .stSelectbox [data-baseweb="select"] {{
      background:{TR_CARD2} !important; border:1px solid {TR_TextArea textarea:focus {{ border-color:{TR_ORANGE} !important; box-shadow:0 0 0 2px rgba(255,128,0,.BORDER} !important; border-radius:6px !important;
  }}
  .stTabs [data-baseweb="tab-list"] {{
      background:{TR_CARD}2) !important; }}
  .stMultiSelect [data-baseweb="select"], .stSelectbox [data-baseweb !important; border-radius:8px 8px 0 0 !important;
      border-bottom:2px solid {TR_ORANGE="select"] {{
      background:{TR_CARD2} !important; border:1px solid {TR_BORDER} !important; border-radius:6px !important;
  }}
  .stT} !important; gap:4px !important;
  }}
  .stTabs [data-baseweb="tab"] {{ colorabs [data-baseweb="tab-list"] {{
      background:{TR_CARD} !important; border-radius:8px 8:{TR_TEXT_MUTED} !important; font-weight:600 !important; borderpx 0 0 !important;
      border-bottom:2px solid {TR_ORANGE} !important; gap:4px !important;
  }}-radius:6px 6px 0 0 !important; }}
  .stTabs [aria-selected="true"] {{ background:{TR_ORANGE} !important; color:#fff !important
  .stTabs [data-baseweb="tab"] {{ color:{TR_TEXT_MUTED} !important; font; }}
  .streamlit-expanderHeader {{
      background:{TR_CARD2} !important; border:1px solid {TR_BORDER} !important;-weight:600 !important; border-radius:6px 6px 0 0 !important; }}
      border-radius:6px !important; color:{TR_TEXT} !important;
  }}
  .stProgress >
  .stTabs [aria-selected="true"] {{ background:{TR_ORANGE} !important; color:#fff !important; }}
  .streamlit-expanderHeader {{
      background:{TR_ div > div {{ background:{TR_ORANGE} !important; }}
  .result-grid {{
      display:grid; gridCARD2} !important; border:1px solid {TR_BORDER} !important;
      border-radius:6px !important; color:{TR_TEXT} !important-template-columns:repeat(auto-fill,minmax(200px,1fr)); gap:10px; margin-top:12px;
  }}
  .result-item;
  }}
  .stProgress > div > div {{ background:{TR_ORANGE} !important; }} {{ background:{TR_CARD2}; border:1px solid {TR_BORDER}; border-radius:8px; padding:10px 
  .result-grid {{
      display:grid; grid-template-columns:repeat(auto-fill,minmax(200px,1fr)); gap:1014px; }}
  .result-item-label {{ font-size:10px; color:{TR_TEXT_MUTED}; text-transform:uppercase; letter-spacing:.5px; }}px; margin-top:12px;
  }}
  .result-item {{ background:{TR_CARD2}; border:1px solid {TR_BORDER}; border
  .result-item-value {{ font-size:16px; font-weight:700; color:{TR_TEXT}; margin-top:2px; }}
  .result-item-value.highlight-radius:8px; padding:10px 14px; }}
  .result-item-label {{ font-size:10px; color:{TR_TEXT {{ color:{TR_ORANGE}; }}
  .cnpj-preview-box {{
      background:{TR_CARD2}_MUTED}; text-transform:uppercase; letter-spacing:.5px; }}
  .result-item-value {{ font-size:16px; font-weight:700; color:{TR_TEXT}; margin; border:1px solid {TR_BORDER}; border-radius:8px; padding:12px 16px; margin-top:10px;
  }}
  -top:2px; }}
  .result-item-value.highlight {{ color:{TR_ORANGE}; }}
  .cnpj-preview-.cnpj-chip {{
      display:inline-block; background:#2a1500; border:1px solid {TR_box {{
      background:{TR_CARD2}; border:1px solid {TR_BORDER}; border-radius:8px; padding:12px 16ORANGE};
      color:{TR_TEXT}; border-radius:6px; padding:3px 10px;px; margin-top:10px;
  }}
  .cnpj-chip {{
      display:inline-block; background:#2
      font-size:12px; font-family:'Courier New',monospace; margin:3px;
  }}
  .cnpj-chip.a1500; border:1px solid {TR_ORANGE};
      color:{TR_TEXT}; borderinvalido {{ border-color:{TR_ERROR}; color:{TR_ERROR}; background:#2b0-radius:6px; padding:3px 10px;
      font-size:12px; font-family:'Courier New',monospace;d0d; }}
  .tipo-row {{
      background:{TR_CARD2}; border:1px solid { margin:3px;
  }}
  .cnpj-chip.invalido {{ border-color:{TR_ERROR}; color:{TR_ERROR}TR_BORDER}; border-radius:8px; padding:12px 14px; margin-bottom:8px;
  }}
  ; background:#2b0d0d; }}
  .tipo-row.tr-footer {{
      text-align:center; color:{TR_TEXT_MUTED}; font-size:11px;
      margin {{
      background:{TR_CARD2}; border:1px solid {TR_BORDER}; border-radius:8px; padding:12px 14-top:40px; padding:16px; border-top:1px solid {TR_BORDER};
  }}
  .tr-footer span {{ color:{TR_px; margin-bottom:8px;
  }}
  .tr-footer {{
      text-align:center; color:{TR_TEXT_MUTED}; font-size:ORANGE}; font-weight:700; }}
  ::-webkit-scrollbar {{ width:6px; height:6px; }}
  ::-webkit-scrollbar-track {{ background:{11px;
      margin-top:40px; padding:16px; border-top:1px solid {TR_BORDER};
  }}
  .tr-footer spanTR_BG}; }}
  ::-webkit-scrollbar-thumb {{ background:{TR_ORANGE}; border-radius:3px; }}
  #MainMenu {{ visibility {{ color:{TR_ORANGE}; font-weight:700; }}
  ::-webkit-scrollbar {{ width:6px; height:6px; }}
  ::-webkit-scrollbar-track {{ background:{: hidden; height: 0; }}
  footer {{ visibility: hidden; height: 0; }}
  header {{ visibility: hidden; height:TR_BG}; }}
  ::-webkit-scrollbar-thumb {{ background:{TR_ORANGE}; border-radius:3px; }}
  #MainMenu {{ visibility: hidden; height: 0; }}
  footer {{ visibility: hidden; height: 0; }}
  header 0; }}
  .block-container {{ padding-top:1rem !important; }}
</style>

<script>
  (function() {{
    function abr {{ visibility: hidden; height: 0; }}
  .block-container {{ padding-top:1rem !important; }}irSidebar() {{
      try {{
        var doc = window.parent.document;
        var sidebar = doc.querySelector('
</style>

<script>
  (function() {{
    function abrsection[data-testid="stSidebar"]');
        if (!sidebar) return;
        if (sidebar.getBoundingClientRect().width < 50irSidebar() {{
      try {{
        var doc = window.parent.document;
        var sidebar = doc.querySelector(') {{
          var btn = doc.querySelector('[data-testid="collapsedControl"] button');
          if (btn) btn.click();
        section[data-testid="stSidebar"]');
        if (!sidebar) return;
        if (sidebar.getBoundingClientRect().width < 50}}
      }} catch(e) {{}}
    }}
    if (document.readyState === 'complete') {{ setTimeout(abrirSidebar) {{
          var btn = doc.querySelector('[data-testid="collapsedControl"] button');
          if (btn) btn.click();
        , 500); }}
    else {{ window.addEventListener('load', function() {{ setTimeout(abrirSidebar, 500); }}); }}
  }})();
</script>
"""}}
      }} catch(e) {{}}
    }}
    if (document.readyState === 'complete') {{ setTimeout(abrirSidebar

TIPOS_EMPRESA = {
    1: "Empresa",
    2: "Tom, 500); }}
    else {{ window.addEventListener('load', function() {{ setTimeout(abrirSidebar, 500); }}); }}
  }})();
</script>
"""ador de Serviço",
    3: "Empreitada Parc

TIPOS_EMPRESA = {
    1: "Empresa",
    2: "Tomador deial",
    4: "Obra Própria",
    5: "Empreitada Total",
    6: " Serviço",
    3: "Empreitada Parcial",
    4: "Cooperativa de Trabalho",
}

# ──────────────────────────────Obra Própria",
    5: "Empreitada Total",
    6: "Cooperativa de────────────────────────────────────────────────
# MUNICÍPIOS
# ──────────────────────────────────────────────────────────────────────────────
def Trabalho",
}

# ──────────────────────────────────────────────────────────────────────────────
# MUNICÍPIOS
# ────── _normalizar(s: str) -> str:
    s = str────────────────────────────────────────────────────────────────────────
def _normalizar(s(s).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicode: str) -> str:
    s = str(s).strip().upper()
    s = ''.join(c for c indata.category(c) != 'Mn')
    s = re.sub(r"['\-]", " ", s)
    s = unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re re.sub(r"\s+", " ", s).strip()
    return s

def _carregar_municipios():
    try:
        df = pd.read_excel.sub(r"['\-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def(
            "RELAÇÃO DE MUNICÍPIOS.xlsx",
            sheet_name=" _carregar_municipios():
    try:
        df = pd.read_excel(
            "RELRELAÇÃO DE MUNICÍPIOS",
            engine="openpyxl", dtype=str, headerAÇÃO DE MUNICÍPIOS.xlsx",
            sheet_name="RELAÇÃO DE MUNICÍPIOS",
            engine=0,
        )
        df.columns = [
            unicodedata.normalize('NFC', str="openpyxl", dtype=str, header=0,
        (c)).strip().lstrip('\ufeff').replace('\xa)
        df.columns = [
            unicodedata.normalize('NFC', str(c)).strip().l0',' ').replace('\u200b','')
            for c in df.columns
        ]
        colstrip('\ufeff').replace('\xa0',' ').replace_codigo = col_nome = col_uf = None
        for col in df.columns:
            col_n('\u200b','')
            for c in df.columns
        ]
        col_codigo = col_nome = col = _normalizar(col)
            if col_n in ("CODIGO","COD","CODIGO MUNICIP_uf = None
        for col in df.columns:
            col_n = _normalizar(col)
            if col_n inIO","CODIGO_MUNICIPIO") and col_codigo is None: col_codigo = col
            elif col_n in ("NOME ("CODIGO","COD","CODIGO MUNICIPIO","CODIGO_MUNICIPIO") and col_codigo","MUNICIPIO","NOME MUNICIPIO") and col_nome is None: col_nome = col
            elif col_n in ("ESTADO","UF"," is None: col_codigo = col
            elif col_n in ("NOME","MUNICIPIO","NOME MUNICIPIO") and col_nome is None: col_nome =SIGLA","SIGLA UF") and col_uf is None: col_uf = col
        if col_codigo is None and len(df. col
            elif col_n in ("ESTADO","UF","SIGLA","SIGLA Ucolumns)>=1: col_codigo=df.columns[0]
        if col_nome   is None and len(df.F") and col_uf is None: col_uf = col
        if col_codigo is None and len(df.columns)>=columns)>=2: col_nome=df.columns[1]
        if col_uf     is None and len(df.columns)>=5: col_uf=df.columns[4]
        1: col_codigo=df.columns[0]
        if col_nome   is None and len(df.columns)>=2: col_nome=df.columns[1]
        if col_elif col_uf   is None and len(df.columns)>=3: col_uf=df.columns[2]
        mapauf     is None and len(df.columns)>=5: col_uf=df.columns[4]
        elif col_uf   is None and len(df.columns), erros = {}, []
        for idx, row in df.iterrows():
            try:
                uf_>=3: col_uf=df.columns[2]
        mapa, erros = {raw=str(row[col_uf]).strip(); nome_raw=str(row[col_nome]).strip(); cod}, []
        for idx, row in df.iterrows():
            try:
                uf_raw=str_raw=str(row[col_codigo]).strip()
                if cod_raw.lower() in ("nan","none","","(row[col_uf]).strip(); nome_raw=str(row[col_nome]).strip(); cod_raw=str(row[col_codigo]).strip()0"): continue
                if uf_raw.lower()  in ("nan","none",""): continue
                if
                if cod_raw.lower() in ("nan","none","","0" nome_raw.lower() in ("nan","none",""): continue
                if "." in cod_raw:
                    try): continue
                if uf_raw.lower()  in ("nan","none",""): continue
                if nome_raw.lower() in ("nan","none","": cod_raw=str(int(float(cod_raw)))
                    except: pass
                uf_n): continue
                if "." in cod_raw:
                    try: cod_raw=str(int=_normalizar(uf_raw); nome_n=_normalizar(nome_raw)
                if uf_n and nome_n and cod(float(cod_raw)))
                    except: pass
                uf_n=_normalizar(uf__raw: mapa[(uf_n,nome_n)]=cod_raw
            except Exception as e: erros.append(fraw); nome_n=_normalizar(nome_raw)
                if uf_n and nome_n and cod_raw: mapa[("Linha {idx}: {e}")
        debug={"total":len(mapa),"uf_n,nome_n)]=cod_raw
            except Exception as e: erros.append(f"Linhacol_codigo":col_codigo,"col_nome":col_nome,"col_uf":col_uf,
               "colunas":list(df.columns),"a {idx}: {e}")
        debug={"total":len(mapa),"col_codigo":col_codigo,"mostra":[f"({u},{n})→col_nome":col_nome,"col_uf":col_uf,
               "colunas":list(df.columns),"amostra"{c}" for (u,n),c in list(mapa.items())[:8]],"erros":erros[:5]}
        return m:[f"({u},{n})→{c}" for (u,n),c in list(mapa.items())[:8apa, debug
    except Exception as e:
        return {}, {"erro_fatal":str(e),"total":0}]],"erros":erros[:5]}
        return mapa, debug
    except Exception as e:
        return {}, {"erro

if "MUNICIPIOS_MAP" not in st.session_state:
    _mapa, _debug = _carr_fatal":str(e),"total":0}

if "MUNICIPIOS_MAPegar_municipios()
    st.session_state["MUNICIPIOS_MAP"] = _mapa
    st.session_state["_mun_debug"]" not in st.session_state:
    _mapa, _debug = _carregar_municipios()
    st.session_state["MUNICIPIOS_MAP"] = _mapa
    st     = _debug

def buscar_codigo_municipio(municipio: str, .session_state["_mun_debug"]     = _debug

def buscar_codigo_uf: str) -> str:
    if not municipio or not uf: return ""
    mapa = st.session_state.get("MUNICIPIOS_MAP", {})
    if not mapa: return ""
    municipio(municipio: str, uf: str) -> str:
    if not municipio or not uf: return ""
    mapa = st.session_state.getuf_n=_normalizar(uf); nome_orig=_normalizar(municipio)
    if not uf_n or not nome_orig: return ""
    cod("MUNICIPIOS_MAP", {})
    if not mapa: return ""
    uf_n=_normalizar(uf); nome_orig=_normalizar(municipio)
    if not=mapa.get((uf_n,nome_orig))
    if cod: return str(cod)
    variacoes=set uf_n or not nome_orig: return ""
    cod=mapa.get((uf_n,nome_orig))
    if cod: return str(cod)
    var()
    variacoes.add(re.sub(r"^(DO|DE|DAiacoes=set()
    variacoes.add(re.sub(r"^(|DOS|DAS)\s+","",nome_orig).strip())
    variacoesDO|DE|DA|DOS|DAS)\s+","",nome_orig)..add(re.sub(r"\s+(DO|DE|DA|DOS|DAS)\s+"," ",nome_orig).strip())
    variacoes.addstrip())
    variacoes.add(re.sub(r"\s+(DO|DE|DA|DOS|DAS)\s+"," ",nome_orig).strip())(nome_orig.replace("STO ","SANTO ").replace("S
    variacoes.add(nome_orig.replace("STO ","SANTOTA ","SANTA "))
    variacoes.add(nome_orig.replace("SANTO ","STO ").replace("STA ","SANTA "))
    variacoes.add(nome_orig.replace("SANTO ","S ").replace("SANTA ","STA "))
    variacoes.add(re.sub(r"['\-]"," ",nome_orig).strip())
    variacoes.TO ").replace("SANTA ","STA "))
    variacoes.add(re.sub(r"['\-]"," ",nome_orig).stripadd(nome_orig.replace("D ","DE ").replace("D'","DE "))
    variacoes.dis())
    variacoes.add(nome_orig.replace("D ","DE ").replace("D'","DE card(nome_orig)
    for v in variacoes:
        if not v: continue
        cod=mapa.get((uf_n,v))
        if cod: return str("))
    variacoes.discard(nome_orig)
    for v in variacoes:
        if not v: continue
        codcod)
    for (u,n),c in mapa.items():
        if u!=uf_n: continue
        if nome=mapa.get((uf_n,v))
        if cod: return str(cod)
    for (u,n),c in mapa.items():
        if u_orig in n or n in nome_orig: return str(c)
    prefixo=nome_orig[:!=uf_n: continue
        if nome_orig in n or n in nome_orig: return str5]
    if len(prefixo)>=4:
        for (u,n),c in m(c)
    prefixo=nome_orig[:5]
    if len(prefixo)>=4:
        forapa.items():
            if u==uf_n and n.startswith(prefixo): return str(c)
    return ""

# ────────────────────────────────────────────────── (u,n),c in mapa.items():
            if u==uf_n and n.startswith(prefixo): return str(c)
    return ""

#────────────────────────────
# MOTOR FPAS
# ────────────────────────────────────────────────────────────────────────────── ──────────────────────────────────────────────────────────────────────────────
# MOTOR FPAS
# ──────────────────────────────────────────────────────────────────
ENTIDADES = {
    "FNDE":    {"codigo────────────
ENTIDADES = {
    "FNDE":    {"":1,    "aliquota":2.codigo":1,    "aliquota":2.5,  "nome":"Salário-Educação"},
    "INCRA":   {"codigo":2,    "aliquota":5,  "nome":"Salário-Educação"},
    "INCRA":   {"codigo":2,    "aliquota":0.2,  "nome":"INCRA"},
    "SENAI":   {"codigo":4,    "aliquota":0.2,  "nome":"INCRA"},
    "SENAI":   {"codigo":4,    "aliquota":1.0,  "nome":"SENAI"},
    "SESI":    {"codigo":8,    "aliquota":11.0,  "nome":"SENAI"},
    "SESI":    {"codigo":8,    "aliquota":1.5,  "nome":"SESI"},
    "SENAC":   {"codigo":16,   "aliquota":1.0,  "nome":"SENAC"},
    "SESC":    {"codigo":32,   "aliquota":1.5,  "nome":"SESC"},
    "SEBRAE":  {"codigo":64,   "aliquota":0.6,  "nome":"SEBRAE"},
    "SENAR":   {"codigo":128,  "aliquota":2.5,  "nome":"SENAR"},
    "SEST":    {"codigo":256,  "aliquota":1.5,  "nome":"SEST"},
    "SENAT":   {"codigo":512,  "aliquota":1.0,  "nome":"SENAT"},
    "SESCOOP": {"codigo":1024,.5,  "nome":"SESI"},
    "SENAC":   {"codigo":16,   "aliquota":1.0,  "nome":"SENAC"},
    "SESC":    {"codigo":32,   "aliquota":1.5,  "nome":"SESC"},
    "SEBRAE":  {"codigo":64,   "aliquota":0.6,  "nome":"SEBRAE"},
    "SENAR":   {"codigo":128,  "aliquota":2.5,  "nome":"SENAR"},
    "SEST":    {"codigo":256,  "aliquota":1.5,  "nome":"SEST"},
    "SENAT":   {"codigo":512,  "aliquota":1.0,  "nome":"SENAT"},
    "SESCOOP": {"codigo":1024, "aliquota":2.5,  "nome":"SESCOOP"},
}

FPAS_CONFIG = {
    507 "aliquota":2.5,  "nome":"SESCOOP"},
}

FPAS_CONFIG = {
    507:(79,  :(79,  2,"2100","2,"2100","150","Ind150","Indústria Geral"),
    515:(115, 2,"2100","ústria Geral"),
    515:(115, 2,"2100","150","Comércio / Serviços em150","Comércio / Serviços em Geral"),
    566:(99,  2,"2100","150","Prof Geral"),
    566:(99,  2,"2100","150","Profissional Liberal / Autônomo Pissional Liberal / Autônomo PF"),
    574:(115F"),
    574:(115, 1,"2100","150","Saúde"),
    582:(0,   1,"2100","150","Administ, 1,"2100","150","Saúde"),
    582:(0,   1,"2100","150","Administraçãoração Pública"),
    604:(3,   2,"2631","150","Produtor Rural P Pública"),
    604:(3,   2,"2631","150","Produtor Rural PF"),
    620F"),
    620:(145, 2,"2100","150","Transporte Rodoviário / A:(145, 2,"2100","150","Transporte Rodoviário / Aéreo"),
    639éreo"),
    639:(0,   0,"2100","150","Entidade Beneficente:(0,   0,"2100","150","Entidade Beneficente Isenta Isenta"),
    647:(115, 1,"2100","150","Educação /"),
    647:(115, 1,"2100","150","Educação / Assistência Social Assistência Social"),
    655:(115, 2,"2100","150","Instituição Financ"),
    655:(115, 2,"2100","150","Instituição Financeira"),
    663:(115, 2eira"),
    663:(115, 2,"2100","150","Comunicação / Publicidade"),
    671:(115, 2,"2100","150","Hot,"2100","150","Comunicação / Publicidade"),
    671:(115, 2,"2100","150","Hotéis / Turismo"),
    680éis / Turismo"),
    680:(115, 2,"2100","150","Limpeza / Conservação"),
    736:(115, 2,"2100","150","T:(115, 2,"2100","150","Limpeza / Conservação"),
    736:(115, 2,"2100","150","TI / Tecnologia da Informação"),
    744I / Tecnologia da Informação"),
    744:(115, 2,"2100","150","Consultoria / Assessoria"),
    760:(115, 2,"2100","150","Seg:(115, 2,"2100","150","Consultoria / Assessoria"),
    760:(115, 2,"2100","150","Segurança Privurança Privada"),
    779:(115, 2,"2100","150","Serviços Pessoais"),
    787:(3ada"),
    779:(115, 2,"2100","150","Serviços Pessoais"),
    787:(3,   2,"2631","150","Produtor Rural P,   2,"2631","150","Produtor Rural PJ"),
    795:(115, 2,"2100","150","Cooperativa de Trabalho"),
    803:(115, 2J"),
    795:(115, 2,"2100","150","Cooperativa de Trabalho"),
    803:(115, 2,"2100","150","Cooper,"2100","150","Cooperativa de Produção"),
    810:(1024,2,"ativa de Produção"),
    810:(1024,2,"2100","150","Cooperativa de Cr2100","150","Cooperativa de Crédito (SESCOOP)"),
    833:(115, 3,"2100","150","édito (SESCOOP)"),
    833:(115, 3,"2100","150","Construção CivilConstrução Civil"),
    868:(0,   0,"2100","150","Missão Diplomática /"),
    868:(0,   0,"2100","150","Missão Diplomática / Org Org. Internacional"),
    876:(0,   0,"2100","150","Doméstico /. Internacional"),
    876:(0,   0,"2100","150","Doméstico / Contribu Contribuinte Individual"),
}

FPAS_SEMPRE_ZERO = {582, 868inte Individual"),
}

FPAS_SEMPRE_ZERO = {582, 868, 876}

CNAE_FPAS = {
    "0111-3, 876}

CNAE_FPAS = {
    "0111-/01":604,"0111-3/02":604,"0111-3/03":604,"3/01":604,"0111-3/02":6040112-1/01":604,"0113-0/00":604,
    "0115-6/00":604,"0116-4/00":604,"0119-9/01":604,"0119-9/99,"0111-3/03":604,"0112-1/01":604,"0113-0/00":604,
    "0115-6/00":604,"0116-4/00":604,"0119-9/01":604,"0119-9":604,"0121-1/01":604,
    "0122-9/00":604,"0131-8/00":604,"0132-6/99":604,"0121-1/01":604,
    "0122-9/00":604,"0131-8/00":604,"0132/00":604,"0133-4/01":604,"0134-2/00":604,
    "0135-1/00":604,"0139-3/01-6/00":604,"0133-4/01":604,"0134-2/00":604,
    "0135-1/00":604,"0139-3":604,"0139-3/99":604,"0141-5/01":604,"0142-3/00":604,
    "0151-2/01":604,"0139-3/99":604,"0141-5/01":604,"0142-3/00":604,
    "0151/01":604,"0151-2/02":604,"0152-1/01":604,"0153-9/01-2/01":604,"0151-2/02":604,"0152-1/01":604,"0153-9":604,"0154-7/00":604,
    "0155-5/01":604,"0155-5/02":604,"0159-8/01":604,"0154-7/00":604,
    "0155-5/01":604,"0155-5/02":604,"0159/01":604,"0159-8/99":604,"0161-0/00":604,
    "0162-8/00":604,"0163-6-8/01":604,"0159-8/99":604,"0161-0/00":604,
    "0162-8/00":604,"0163-6/00":604,"0170-9/00":604,"0210-1/01":604,"0210-1/02":604,
    "0220/00":604,"0170-9/00":604,"0210-1/01":604,"0210-1/02":604,
    "0220-9/01":604,"0220-9/02":604,"0230-6/00":604,"0240-3/00":604,"0311-9/01":604,"0220-9/02":604,"0230-6/00":604,"0240-3/00-6/01":604,
    "0312-4/00":604,"0321-3/01":604,"0311-6/01":604,
    "0312-4/00":604,"0321-3":604,"0322-1/00":604,
    "1011-2/01":507/01":604,"0322-1/00":604,
    "1011-2/01,"1011-2/02":507,"1012-1/01":507,"1013-9/01":507,"1020-1/01":507,
    "1020":507,"1011-2/02":507,"1012-1/01":507,"1013-9/01":507,"1020-1/01":507,-1/02":507,"1031-7/00":507,"1032-5/01":507,"1033-3/01":507,"
    "1020-1/02":507,"1031-7/00":507,"1032-5/01":507,"10331041-4/00":507,
    "1042-2/00":507,"1043-1/00":507,"1051-1/00":507,"-3/01":507,"1041-4/00":507,
    "1042-2/00":507,"1043-1/00":507,"10511052-0/00":507,"1061-9/01":507,
    "1062-7/00":507,"1063-5/00-1/00":507,"1052-0/00":507,"1061-9/01":507,
    "1062-7/00":507,"1063":507,"1064-3/00":507,"1065-1/01":507,"1066-0/00":507,
    "1069-4/00":507,"1071-5/00":507,"1064-3/00":507,"1065-1/01":507,"1066-0/00":507,
    "1069-4-6/00":507,"1072-4/01":507,"1073-2/01":507,"1081-3/01/00":507,"1071-6/00":507,"1072-4/01":507,"1073-2/01":507,"1081":507,
    "1082-1/00":507,"1083-0/00":507,"1084-8/00":507,"1085-6/00":507,"1086-4/00":-3/01":507,
    "1082-1/00":507,"1083-0/00":507,"1084-8/00":507,"1085507,
    "1087-2/00":507,"1088-1/00":507,"1089-9/00":507,"1091-1/01-6/00":507,"1086-4/00":507,
    "1087-2/00":507,"1088-1/00":507,"1089-9/00":507,"1091-1/02":507,
    "1092-9/00":507,"1093-7/01":507,"1094-5/00":507,"1091-1/01":507,"1091-1/02":507,
    "1092-9/00":507,"1093-7/01":507,"1095-3/00":507,"1096-1/00":507,
    "1099-6/01":507,"1099":507,"1094-5/00":507,"1095-3/00":507,"1096-1/00":507,
    "1099-6-6/99":507,"1111-9/01":507,"1111-9/02":507,"1112-7/00":507,
    "1113-5/01/01":507,"1099-6/99":507,"1111-9/01":507,"1111-9/02":507,"1112-7":507,"1121-6/00":507,"1122-4/01":507,"1122-4/03":507,"1210-7/00":507,
    "1113-5/01":507,"1121-6/00":507,"1122-4/01":507,"1122-4/03/00":507,
    "1220-4/01":507,"1311-1/00":507,"1312-0":507,"1210-7/00":507,
    "1220-4/01":507,"1311-1/00":507,"1313-8/00":507,"1314-6/00":507,
    "1321-9/00":507,"1322-7/00":507,"1323-5/00":507,"1330-8/00":507,"1340-5/01":507,
    "1351/00":507,"1312-0/00":507,"1313-8/00":507,"1314-6/00":507,
    "1321-9/00":507,"1322-7/00":507,"1323-5/00":507,"1330-8/00":507,"1340-5/01-1/00":507,"1352-9/00":507,"1353-7/00":507,"1354-5/00":507,"1359-6":507,
    "1351-1/00":507,"1352-9/00":507,"1353-7/00":507,"1354-5/00":/00":507,
    "1411-8/01":507,"1411-8/02":507,"1412-6/01":507,"1413-4507,"1359-6/00":507,
    "1411-8/01":507,"1411-8/02":507,"/01":507,"1414-2/00":507,
    "1421-5/00":507,"1422-3/00":507,"15101412-6/01":507,"1413-4/01":507,"1414-2/00":507,
    "1421-5/00":507,"1422-6/00":507,"1521-1/00":507,"1529-7/00":507,
    "1531-9/01-3/00":507,"1510-6/00":507,"1521-1/00":507,"1529-7/00":507,
    ":507,"1532-7/00":507,"1533-5/00":507,"1539-4/00":507,"1540-8/00":507,
    "1531-9/01":507,"1532-7/00":507,"1533-5/00":507,"1539-4/00":507,"1540"1610-2/01":507,"1610-2/02":507,"1621-8/00":507,"1622-6/01-8/00":507,
    "1610-2/01":507,"1610-2/02":507,"1621":507,"1710-9/00":507,
    "1721-4/00":507,"1722-2/00":507,"1731-8/00":507,"1622-6/01":507,"1710-9/00":507,
    "1721-4-1/00":507,"1732-0/00":507,"1733-8/00":507,
    "1741-9/01":507,"1742-7/00":507,"1722-2/00":507,"1731-1/00":507,"1732-0/00":507,"1733-8/00":507,/00":507,"1749-4/00":507,"1811-3/01":507,"1811-3/02":507,
    "1812-1/00
    "1741-9/01":507,"1742-7/00":507,"1749-4/00":507,"1811-3":507,"1813-0/00":507,"1821-1/00":507,"1822-9/01":507,"1830-0/01":507,"1811-3/02":507,
    "1812-1/00":507,"1813-0/00":507,"1821-1/00":507,"/01":507,
    "1830-0/02":507,"1830-0/03":507,"1910-1/001822-9/01":507,"1830-0/01":507,
    "1830-0":507,"1921-7/00":507,"1922-5/00":507,
    "1931-4/00":507,"1932/02":507,"1830-0/03":507,"1910-1/00":507,"1921-7/00":507,"1922-5-2/00":507,"1933-1/00":507,"2011-8/00":507,"2012-6/00":507,
    "2013-4/00":507,
    "1931-4/00":507,"1932-2/00":507,"1933-1/00":507,"2011-8/01":507,"2021-5/00":507,"2022-3/00":507,"2023-1/00":507,"2029/00":507,"2012-6/00":507,
    "2013-4/01":507,"2021-5/00":507,"2022-3/00":507,"-1/00":507,
    "2031-2/00":507,"2032-1/00":507,"2033-92023-1/00":507,"2029-1/00":507,
    "2031-2/00":507,"2032-1/00":507,"2040-1/00":507,"2051-7/00":507,
    "2052-5/00":507,"2061-4/00":507,"2033-9/00":507,"2040-1/00":507,"2051-7/00":507,
    "2052-5/00/00":507,"2062-2/00":507,"2063-1/00":507,"2064-9/00":507,
    "2065-7/00":507,"2066-5/00":507,"":507,"2061-4/00":507,"2062-2/00":507,"2063-1/00":507,"2064-9/00":507,
    "2065-7/00":2067-3/00":507,"2068-1/00":507,"2091-6/00":507,
    "2092-4/01":507,"2093-2507,"2066-5/00":507,"2067-3/00":507,"2068-1/00":507,"2091-6/00":507,
    /00":507,"2094-1/00":507,"2099-1/01":507,"2099-1/99":507,
    "2"2092-4/01":507,"2093-2/00":507,"2094-1/00":507,"2099-1/01":507,"2099110-6/00":507,"2121-1/01":507,"2121-1/02":507,"2122-0-1/99":507,
    "2110-6/00":507,"2121-1/01":507,"2121-1/02":507,"/00":507,"2123-8/00":507,
    "2211-1/00":507,"2212-9/00":507,"2219-6/00":507,"22212122-0/00":507,"2123-8/00":507,
    "2211-1/00":507,"2212-9/00":507,"2219-6-8/00":507,"2222-6/00":507,
    "2223-4/00":507,"2229-3/01":507,"2311/00":507,"2221-8/00":507,"2222-6/00":507,
    "2223-4/00":507,"2229-3/01-7/00":507,"2312-5/00":507,"2319-2/00":507,
    "2320-6":507,"2311-7/00":507,"2312-5/00":507,"2319-2/00":507,
    "2320-6/00":507,"2330-3/01":507,"2341-9/00":507,"2342-7/00":507,"2349/00":507,"2330-3/01":507,"2341-9/00":507,"2342-7/00":507,"2349-4/00":507,
    "2350-8/00":507,"2391-5/00":507,"2392-3/00":507,"2399-1/01":507,"2410-1/00":507,
    "2421-1/00":507,"2422-9/01":507,"2423-7/01":507,"2424-5/00":507,"2431-8/00":507,
    "2439-3-4/00":507,
    "2350-8/00":507,"2391-5/00":507,"2392-3/00":507,"2399-1/01":507,"2410-1/00":507,
    "2421-1/00":507,"2422-9/01":507,"2423-7/01":507,"2424-5/00":507,"2431-8/00":507,
    "2439-3/00/00":507,"2441-5/01":507,"2442-3/00":507,"2443-1/00":507,"2444-0":507,"2441-5/01":507,"2442-3/00":507,"2443-1/00":507,"2444-0/00":507,
    /00":507,
    "2445-8/00":507,"2446-6/00":507,"2449-1/99"2445-8/00":507,"2446-6/00":507,"2449-1/99":507,"2451-2":507,"2451-2/00":507,"2452-1/00":507,
    "2511-0/00":507,"2512-8/00":507,"2452-1/00":507,
    "2511-0/00":507,"2512-8/00":507,"2513/00":507,"2513-6/00":507,"2521-7/00":507,"2522-5/00":507,
    "2531-4/01-6/00":507,"2521-7/00":507,"2522-5/00":507,
    "2531-4/01":507,"2532":507,"2532-2/00":507,"2539-0/01":507,"2541-1/00":507,"2542-0-2/00":507,"2539-0/01":507,"2541-1/00":507,"2542-0/00":507,
    /00":507,
    "2543-8/00":507,"2550-1/00":507,"2591-8/00":507,"2592-6"2543-8/00":507,"2550-1/00":507,"2591-8/00":507,"2592-6/01":507,"2593-4/01":507,"2593-4/00":507,
    "2599-3/01":507,"2599/00":507,
    "2599-3/01":507,"2599-3/99":507,"2610-3/99":507,"2610-8/00":507,"2621-3/00":507,"2622-1/00":507,
    "2631-8/00":507,"2621-3/00":507,"2622-1/00":507,
    "2631-1/00":507,"2632-1/00":507,"2632-9/00":507,"2640-0/00":507,"2651-5/00":507,"2652-9/00":507,"2640-0/00":507,"2651-5/00":507,"2652-3/00":507,
    "2660-3/00":507,
    "2660-4/00":507,"2670-1/01":507,"2680-9-4/00":507,"2670-1/01":507,"2680-9/00":507,"2710-4/00":507,"2710-4/01":507,"2710-4/02":507,
    "2721-0/01":507,"2710-4/02":507,
    "2721-0/00":507,"2722-8/00":507,"/00":507,"2722-8/00":507,"2731-7/00":507,"2732-5/00":507,"2733-3/00":507,
    "27402731-7/00":507,"2732-5/00":507,"2733-3/00":507,
    "2740-6-6/01":507,"2751-1/00":507,"2759-7/01":507,"2790/01":507,"2751-1/00":507,"2759-7/01":507,"2790-2-2/01":507,"2811-9/00":507,
    "2812-7/00":507,"2813-5/00":507,"2814-3/00":507,"2815-1/00":507,"2821-6/00":507,
    "2822-4/01/01":507,"2811-9/00":507,
    "2812-7/00":507,"2813-5/00":507,"2814-3/00":507,"2815-1/00":507,"2821-6/00":507,
    "2822-4/01":507,"2823-2":507,"2823-2/00":507,"2824-1/01":507,"2825-9/00":507,"2829-1/00":507,"2824-1/01":507,"2825-9/00":507,"2829-1/01/01":507,
    "2830-5/00":507,"2840-2/00":507,"2851-8":507,
    "2830-5/00":507,"2840-2/00":507,"2851-8/00":507,"2852-6/00":507,"/00":507,"2852-6/00":507,"2853-4/00":507,
    "2854-2/00":507,"2861-5/00":507,"2862-3/00":507,"2863-1/00":507,"2864-0/00":507,
    "2865-8/00":507,"2866-6/00":507,"2869-1/00":507,"2853-4/00":507,
    "2854-2/00":507,"2861-5/00":507,"2862-3/00":507,"2863-1/00":507,"2864-0/00":507,
    "2865-8/00":507,"2866-6/00":507,"2869-1/00":507,"2910-72910-7/01":507,"2920-4/01":507,
    "2930/01":507,"2920-4/01":507,
    "2930-1/01":507,"2941-1/01":507,"2941-7/00":507,"2942-5/00":507,"2943-3/00":507,"2944-1/00":507,
    "2945-0-7/00":507,"2942-5/00":507,"2943-3/00":507,"2944-1/00":507,
    "2945-0/00":507,"2949-2/00":507,"2949-2/01":507,"2950-6/00":507,"3011-3/01":507,"3012/01":507,"2950-6/00":507,"3011-3/01":507,"3012-1/00":507,-1/00":507,
    "3031-8/00":507,"3032-6/00":507,"3041-5
    "3031-8/00":507,"3032-6/00":507,"3041-5/00":507,"3042-3/00":507,"/00":507,"3042-3/00":507,"3050-4/00":507,
    "3091-1/00":507,"3092-03050-4/00":507,
    "3091-1/00":507,"3092-0/00":507,"3099-7/00":507,"3101/00":507,"3099-7/00":507,"3101-2/00":507,"3102-1/00":507,
    "3103-9/00-2/00":507,"3102-1/00":507,
    "3103-9/00":507,"3104-7/00":507,"3211-6":507,"3104-7/00":507,"3211-6/01":507,"3212-4/00":507,"3220-5/01":507,"3212-4/00":507,"3220-5/00":507,
    "3230-2/00":507,
    "3230-2/00":507,"3240-0/01":507,"3291-4/00":507,"3240-0/01":507,"3291-4/00":507,"3292-2/00":507,"3299-0/00":507,"3292-2/00":507,"3299-0/05":507,
    "4110-7/00":833/05":507,
    "4110-7/00":833,"4120-4/00":833,"4211,"4120-4/00":833,"4211-1/01":833,"4211-1/02":833,"4212-1/01":833,"4211-1/02":833,"4212-0/00":833,
    "4213-0/00":833,
    "4213-8/00":833,"4221-9/01":833,"4221-9/02":833,"4222-7-8/00":833,"4221-9/01":833,"4221-9/02":833,"4222-7/00":833,"4223/00":833,"4223-5/00":833,
    "4291-0/00":833,"4292-8/00":833,"4299-5-5/00":833,
    "4291-0/00":833,"4292-8/00":833,"4299-5/01/01":833,"4299-5/99":833,"4311-8/01":833,
    "4311-8/02":833,"4312":833,"4299-5/99":833,"4311-8/01":833,
    "4311-8/02":833,"4312-6/00":833,"4313-4/00":833,"4319-3/00":833,"4321-5/00":833,
    "4322-3-6/00":833,"4313-4/00":833,"4319-3/00":833,"4321-5/00":833,
    /01":833,"4322-3/02":833,"4323-1/00":833,"4329-1/01":833,"4329-1"4322-3/01":833,"4322-3/02":833,"4323-1/00":833,"4329-1/01/05":833,
    "4329-1/99":833,"4330-4/01":833,"4330-4":833,"4329-1/05":833,
    "4329-1/99":833,"4330-4/01/02":833,"4330-4/03":833,"4330-4/04":833,
    "4330-4/05":833,"4391-6/00":833,"4399":833,"4330-4/02":833,"4330-4/03":833,"4330-4/04":833,
    "4330-4/05":833,"4391-6-1/01":833,"4399-1/02":833,"4399-1/03":833,
    "4399-1/04":833,"4399-1/05":833,"/00":833,"4399-1/01":833,"4399-1/02":833,"4399-1/03":833,
    "4399-1/04":833,"43994399-1/99":833,
    "4511-1/01":515,"4511-1/02":515,"4512-1/05":833,"4399-1/99":833,
    "4511-1/01":515,"4511-1/02-9/01":515,"4512-9/02":515,"4520-0/01":515,
    "4520-0":515,"4512-9/01":515,"4512-9/02":515,"4520-0/01":515,/02":515,"4530-7/01":515,"4530-7/02":515,"4530-7/03":515,"4541
    "4520-0/02":515,"4530-7/01":515,"4530-7/02":515,"4530-7/03":515,"-2/01":515,
    "4541-2/02":515,"4541-2/03":515,"4542-1/004541-2/01":515,
    "4541-2/02":515,"4541-2/03":515,"4542":515,"4543-9/00":515,"4611-7/00":515,
    "4612-5/00":515,"4613-1/00":515,"4543-9/00":515,"4611-7/00":515,
    "4612-5/00":515,"4613-3/00":515,"4614-1/00":515,"4615-0/00":515,"4616-8/00":515,
    "4617-6/00":515,"4618-3/00":515,"4614-1/00":515,"4615-0/00":515,"4616-8/00":515,
    "4617-6/00":515,"4618-4/01":515,"4619-2/00":515,"4621-4/00":515,"4622-2/00":515,
    "4623-1/01":515,"4623-1/08":515,"4631-1/00":515,"4632-0/01":515,"4633-4/01":515,"4619-2/00":515,"4621-4/00":515,"4622-2/00":515,
    "4623-1/01":515,"4623-1/08":515,"4631-1/00":515,"4632-0/01-8/01":515,
    "4634-6/01":515,"4635-4/01":515,"4636-2/01":515,"4637-1/01":515,"4633-8/01":515,
    "4634-6/01":515,"4635-4/01":515,"4639-7/01":515,
    "4641-9/01":515,"4642-7/01":515,"4643-5/01":515,"4644":515,"4636-2/01":515,"4637-1/01":515,"4639-7/01":515,
    "4641-9/01":515,"4642-3/01":515,"4645-1/01":515,
    "4646-0/01":515,"4647-8/01":515,"4649-4-7/01":515,"4643-5/01":515,"4644-3/01":515,"4645-1/01":515,
    "4646-0/01":515,"4651-6/01":515,"4652-4/00":515,
    "4661-3/00":515,"4662-1/01":515,"4647-8/01":515,"4649-4/01":515,"4651-6/01":515,"4652-4/00/00":515,"4663-0/00":515,"4664-8/00":515,"4665-6/00":515,
    "4669-9":515,
    "4661-3/00":515,"4662-1/00":515,"4663-0/00":515,"4664/01":515,"4671-1/00":515,"4672-9/00":515,"4673-7/00":515,"4674-8/00":515,"4665-6/00":515,
    "4669-9/01":515,"4671-1-5/00":515,
    "4679-6/01":515,"4681-8/01":515,"4682-6/00/00":515,"4672-9/00":515,"4673-7/00":515,"4674-5/00":515,
    "4679-6":515,"4683-4/00":515,"4684-2/01":515,
    "4685-1/00":515,"4686-9/00":515,"4687/01":515,"4681-8/01":515,"4682-6/00":515,"4683-4/00":515,"4684-2/01":515,-7/01":515,"4689-3/00":515,"4691-5/00":515,
    "4692-3/00
    "4685-1/00":515,"4686-9/00":515,"4687-7/01":515,"4689-3/00":515,"4693-1/00":515,"4711-3/01":515,"4711-3/02":515,"4712-1/00":515,":515,"4691-5/00":515,
    "4692-3/00":515,"4693-1/00":515,"4711-3/01":515,"4711
    "4713-0/01":515,"4713-0/02":515,"4721-1/01":515,"4721-1/02":515,"4722-9/01":515,-3/02":515,"4712-1/00":515,
    "4713-0/01":515,"4713-0/02":515,"4721-1
    "4723-7/00":515,"4724-5/00":515,"4729-6/01":515,"4731-8/01":515,"4721-1/02":515,"4722-9/01":515,
    "4723-7/00":515,"4724-5/00":515,"/00":515,"4732-6/00":515,
    "4741-5/00":515,"4742-3/00":515,"47434729-6/01":515,"4731-8/00":515,"4732-6/00":515,
    "4741-1/00":515,"4744-0/01":515,"4744-0/02":515,
    "4744-0/03":515,"4744-0/04":515,"-5/00":515,"4742-3/00":515,"4743-1/00":515,"4744-0/01":515,"4744-0/02":515,4744-0/05":515,"4744-0/06":515,"4744-0/99":515,
    "4751-2/01":515,"4751-2/02":515,"
    "4744-0/03":515,"4744-0/04":515,"4744-0/05":515,"4744-0/06":515,"4744-0/994752-1/00":515,"4753-9/00":515,"4754-7/01":515,
    "4755-5/01":515,"4756-3/00":515,
    "4751-2/01":515,"4751-2/02":515,"4752-1/00":515,"4753-9/00":515,"4754":515,"4757-1/00":515,"4759-8/01":515,"4761-0/01":515,-7/01":515,
    "4755-5/01":515,"4756-3/00":515,"4757-1/00":515,"4759-8
    "4762-8/00":515,"4763-6/01":515,"4771-7/01":515,"4771-7/02":515,"4771/01":515,"4761-0/01":515,
    "4762-8/00":515,"4763-6/01":515,"4771-7/03":515,
    "4772-5/00":515,"4773-3/00":515,"4774-1/00":515,"4781-4-7/01":515,"4771-7/02":515,"4771-7/03":515,
    "4772-5/00/00":515,"4782-2/01":515,
    "4783-1/01":515,"4784-9":515,"4773-3/00":515,"4774-1/00":515,"4781-4/00":515,"4782-2/01/00":515,"4785-7/01":515,"4789-0/01":515,"4789-0/99":515,
    "4791-1/00":515,
    "4783-1/01":515,"4784-9/00":515,"4785-7/01":515,"4789":515,"4792-8/00":515,"4793-6/00":515,
    "4911-6/00":620,"4912-4/01-0/01":515,"4789-0/99":515,
    "4791-1/00":515,"4792-8/00":515,"4793":620,"4912-4/02":620,"4921-3/01":620,"4921-3/02":620,
    "4922-1/01-6/00":515,
    "4911-6/00":620,"4912-4/01":620,"4912-4/02":620,"4921-3":620,"4922-1/02":620,"4923-0/01":620,"4923-0/02":620,"4924-8/00/01":620,"4921-3/02":620,
    "4922-1/01":620,"4922-1/02":620,"4923-0/01":620,
    "4929-9/01":620,"4929-9/02":620,"4929-9/99":620,"4923-0/02":620,"4924-8/00":620,
    "4929-9/01":620,"4930-2/01":620,"4930-2/02":620,
    "4930-2/03":620,"4940-0/00":620,"4950-7/00":620,"5011-4/01":620,"5011-4/02":620,
    "5012":620,"4929-9/02":620,"4929-9/99":620,"4930-2/01":620,"4930-2/02":620,
    "4930-2/03":620,"4940-0/00":620,"4950-7/00":620,"5-2/01":620,"5021-1/01":620,"5022-0/01":620,"5030011-4/01":620,"5011-4/02":620,
    "5012-2/01":620,"5021-1-1/00":620,"5091-2/01":620,
    "5099-8/01":620,"5022-0/01":620,"5030-1/00":620,"5091-2/01":620,"5111-1/00":620,"5112-9/00":620,"5120-0/00":620,"5130-7/00":620,/01":620,
    "5099-8/01":620,"5111-1/00":620,"5112-9
    "5141-2/00":620,"5142-1/00":620,"5150-2/00":620,"5161-8/00":620,"5120-0/00":620,"5130-7/00":620,
    "5141-2/00":620,"5162-6/00":620,
    "5163-4/00":620,"5172-3/00":620,"5174-0/00":620,"5179-1/00":620,"5211-7/01":620,
    "5211-7/02":620,"5212/00":620,"5142-1/00":620,"5150-2/00":620,"5161-8/00":620,"5162-6/00":620,
    "5163-4/00":620,"5172-3/00":620,"5174-0/00":620,"5179-1/00":620,"5211-5/00":620,"5221-4/00":620,"5222-2/00":620,"5223-1/00":620,
    "5229-7/01":620,
    "5211-7/02":620,"5212-5/00":620,"5221-0/01":620,"5229-0/02":620,"5229-0/99":620,"5231-1/01":620,"5232-4/00":620,"5222-2/00":620,"5223-1/00":620,
    "5229-0/01-0/00":620,
    "5239-7/00":620,"5240-1/01":620,"5229-0/02":620,"5229-0/99":620,"5231-1/01":620,"5232-0/00":620,"5250-8/01":620,"5250-8/02":620,"5310-5/01":620,
    ":620,
    "5239-7/00":620,"5240-1/01":620,"5250-8"5310-5/02":620,"5320-2/01":620,
    "6201-5/01":620,"5250-8/02":620,"5310-5/01":620,
    "5310-5/02":620,"/00":736,"6201-5/01":736,"6201-5/02":736,"6202-3/00":736,"6203-1/00":736,5320-2/01":620,
    "6201-5/00":736
    "6204-0/00":736,"6209-1/00":736,"6311-9/00":736,"6319-4,"6201-5/01":736,"6201-5/02":736,"6202-3/00":736,"6203-1/00":736,/00":736,"6391-7/00":736,
    "6399-2/00":736,
    "6410-7
    "6204-0/00":736,"6209-1/00":736,"6311-9/00":736,"6319/00":655,"6421-2/00":655,"6422-1/00":655,"6423-9/00":655,"6424-4/00":736,"6391-7/00":736,
    "6399-2/00":736,
    "6410-7/01":655,
    "6424-7/02":655,"6424-7/03":655,"6431-0-7/00":655,"6421-2/00":655,"6422-1/00":655,"6423-9/00":655,"6432-8/00":655,"6433-6/00":655,
    "6434-4/01":655,"6435-2/00":655,"6424-7/01":655,
    "6424-7/02":655,"6424-7/03":655,"6431/01":655,"6436-1/00":655,"6437-9/00":655,"6438-7/00":655,
    "6440-9/00":655,"6450-6/00":655,"6461-1/00":655,"6462-0/00":655,"6463-8/00":655,-0/00":655,"6432-8/00":655,"6433-6/00":655,
    "6434-4/01":655,"6435-2/01":655,"6436-1/00":655,"6437-9/00":655,"6438-7/00":655,
    "6
    "6470-1/01":655,"6491-3/00":655,"6492-1/00":655,"6493-0440-9/00":655,"6450-6/00":655,"6461-1/00":655,"6462-0/00":655,"6499-9/01":655,
    "6499-9/99":655,"6511-1/01":655,"6511-1/02":655,"6512-0/00":655,"6520-1/00":655,
    "6521-9/00":655,"6522-7/00":655,"6523-5/00":655,"6524-3/00":655,"6525-1/00":655,
    "6530-8/00":655,"6541-3/00":655,"6463-8/00":655,
    "6470-1/01":655,"6491-3/00":655,"6492-1/00":655,"6493-0/00":655,"6499-9/01":655,
    "6499-9/99":655,"6511-1/01":655,"6511-1/02":655,"6512-0/00":655,"6520-1/00":655,
    /00":655,"6542-1/00":655,"6550-2/00":655,"6611-8/00":655,
    "6612-6/00"6521-9/00":655,"6522-7/00":655,"6523-5/00":655,"6524-3/00":655,"6525-1/00":655,
    "6530":655,"6613-4/00":655,"6619-3/01":655,"6621-5/00":655,"6622-3-8/00":655,"6541-3/00":655,"6542-1/00":655,"6550-2/00":655,"6611/00":655,
    "6629-1/00":655,"6630-4/00":655,
    "8610-1/01-8/00":655,
    "6612-6/00":655,"6613-4/00":655,"6619-3/01":655,"6621":647,"8610-1/02":647,"8621-6/01":647,"8621-6/02":647,"-5/00":655,"6622-3/00":655,
    "6629-1/00":655,"6630-4/00":655,
    "88622-4/00":647,
    "8630-5/01":647,"8630-5/02":647,"8630610-1/01":647,"8610-1/02":647,"8621-6/01-5/03":647,"8630-5/04":647,"8630-5/06":647,
    "8630-5/07":647,"8630-5/08":":647,"8621-6/02":647,"8622-4/00":647,
    "8630-5/01647,"8640-2/01":647,"8640-2/02":647,"8640-2/03":647,
    "8640-2/04":647,"8640-2/05":":647,"8630-5/02":647,"8630-5/03":647,"8630-5/04":647,"8630-5/06":647,
    "8630-5/07":647647,"8640-2/06":647,"8640-2/07":647,"8640-2/08":647,
    "8640-2/09":647,"8640-2/10":647,"8640-2/11,"8630-5/08":647,"8640-2/01":647,"8640-2/02":647,"8640-2/03":647,
    "8640-2/04":647,"":647,"8640-2/12":647,"8640-2/13":647,
    "8650-0/01":647,"8650-0/02":647,"8650-0/03":8640-2/05":647,"8640-2/06":647,"8640-2/07":647,"8640-2/08":647,
    "8640-2/09":647,"8640-2/10":647647,"8650-0/04":647,"8650-0/05":647,
    "8650-0/06":647,"8650-0/07":647,"8660-7/00":647,
    "8511-2/00":647,"8512-1/00":647,"8513-9/00":647,"8520,"8640-2/11":647,"8640-2/12":647,"8640-2/13":647,
    "8650-0/01":647,"8650-0/02":647,"8650-0/03":647,"8650-0/04":647,"8650-0/05":647,
    "8650-0/06":647,"8650-0/07":647,"8660-1/00":647,"8531-7/00":647,
    "8532-5/00":647,"8533-3/00":647,"8541-4-7/00":647,
    "8511-2/00":647,"8512-1/00":647,"8513-9/00":647,"/00":647,"8542-2/00":647,"8550-3/01":647,
    "8550-3/02":647,"85918520-1/00":647,"8531-7/00":647,
    "8532-5/00":647,"8533-3/00":647,"-1/00":647,"8592-9/01":647,"8593-7/00":647,"8599-68541-4/00":647,"8542-2/00":647,"8550-3/01":647,/01":647,
    "8599-6/99":647,
    "8411-6/00":582
    "8550-3/02":647,"8591-1/00":647,"8592-9/01":647,"8593-7/00":647,"8599,"8412-4/00":582,"8413-2/00":582,"8421-3/00":582,"8422-1/00":582,
    -6/01":647,
    "8599-6/99":647,
    "8411-6"8423-0/00":582,"8424-8/00":582,"8425-6/00":582,"8430-2/00":582,/00":582,"8412-4/00":582,"8413-2/00":582,"8421-3/00":582,"8422-1/00":582,
    "7711-0/00":515,"7719-5/99":515,"7721
    "8423-0/00":582,"8424-8/00":582,"8425-6/00":582,"8430-2/00":582,-7/00":515,"7722-5/00":515,"7723-3/00":515,
    "7729-2/01":515,"7731
    "7711-0/00":515,"7719-5/99":515,"7721-4/00":515,"7732-2/00":515,"7733-1/00":515,"7739-0/01-7/00":515,"7722-5/00":515,"7723-3/00":515,
    "7729-2/01":515,"7731":515,
    "7740-3/00":515,"7810-8/00":515,"7820-5/00":515,"7830-4/00":515,"7732-2/00":515,"7733-1/00":515,"7739-0/01-2/00":515,"7911-2/00":671,
    "7912-1/00":671,"7990-2":515,
    "7740-3/00":515,"7810-8/00":515,"7820-5/00":515,"7830/00":671,"8011-1/01":760,"8011-1/02":760,"8012-9/00":760,
    -2/00":515,"7911-2/00":671,
    "7912-1/00":671,"7990-2"8020-0/01":760,"8021-8/00":760,"8030-7/00":760,"8111-7/00":680/00":671,"8011-1/01":760,"8011-1/02":760,"8012-9/00":760,
    ,"8112-5/00":680,
    "8121-4/00":680,"8122-2/00":680,"8129"8020-0/01":760,"8021-8/00":760,"8030-7/00":760,"8111-7-0/00":680,"8130-3/00":680,"8211-3/00":515,
    "8219-9/00":680,"8112-5/00":680,
    "8121-4/00":680,"8122-2/00":680,"8129/01":515,"8219-9/99":515,"8220-2/00":515,"8230-0/01-0/00":680,"8130-3/00":680,"8211-3/00":515,
    "8219-9":515,"8291-1/00":515,
    "8292-0/00":515,"8299-7/01":515,"8299-7/01":515,"8219-9/99":515,"8220-2/00":515,"8230-0/01":515,"8291/99":515,
}

# ──────────────────────────────────────────────────────────────────────────────
# CLASSIFICAÇÃO-1/00":515,
    "8292-0/00":515,"8299-7/01":515,"8299-7/99":515, FPAS
# ──────────────────────────────────────────────────────────────────────────────
def form
}

# ──────────────────────────────────────────────────────────────────────────────
# CLASSIFICAÇÃO FPAS
# ──────────────────────────────────────────────────────────────────────────────
def formatar_cnae(raw) -> str | None:
    if not raw: return None
    s = str(raw).strip().replace(".","").replace(" ","")
    if "-" in s and "/" in s: return s
    d = re.sub(r"\D","",s)
    if len(d)==7: return f"{d[:4]}-{d[4]}/{d[5:]}"
    return s

def decodificar_terceiros(codigo: int) -> list[dict]:
    return [{"sigla":s,"nome":d["nome"],"aliquota":d["aliquota"]}
            for s,d in ENTIDADES.items()
            if isinstance(codigo,int) and codigo & d["codigoatar_cnae(raw) -> str | None:
    if not raw: return None
    s = str(raw).strip().replace(".","").replace(" ","")
    if "-" in s and "/" in s: return s
    d = re.sub(r"\D","",s)
    if len(d)==7: return f"{d[:4]}-{d[4]}/{d[5:]}"
    return s

def decodificar_terceiros(codigo: int) -> list[dict]:
    return [{"sigla":s,"nome":d["nome"],"aliquota":d["aliquota"]}
            for s,d in ENTIDADES.items()
            if isinstance(codigo,int) and codigo & d["codigo"]]

def classificar(cnae: str, simples: bool=False, fap: float=1.0, convenios: list=None) -> dict:
    convenios = convenios or []
    cnae_"]]

def classificar(cnae: str, simples: bool=False, fapfmt  = formatar_cnae(cnae)
    if not cnae_fmt or cnae_fmt not in CNAE_FP: float=1.0, convenios: list=None) -> dict:
    AS:
        return {"erro": f"CNAE '{cnae}' não mapeado."}
    fpas = CNAE_FPAS[cnae_fmtconvenios = convenios or []
    cnae_fmt  = formatar_cnae(cnae)
    if]
    if fpas not in FPAS_CONFIG:
        return {"erro": f"FPAS {fpas} sem configuração."}
    terc not cnae_fmt or cnae_fmt not in CNAE_FPAS:
        return {"erro": f"CNAE '{cnae}' não meiros_base, rat_base, cod_gpsapeado."}
    fpas = CNAE_FPAS[cnae_fmt]
    if fpas not in FPAS_CONFIG:
        return {"erro":, cod_gfip, descricao = FPAS_CONFIG[fpas]
     f"FPAS {fpas} sem configuração."}
    terceiros_base, ratif simples:
        codigo_terceiro, obs = None, "Simples Nacional — campo terceiros em branco"
    elif fp_base, cod_gps, cod_gfip, descricao = FPAS_CONFIG[fpas]
    if simples:
        codigo_tercas in FPAS_SEMPRE_ZERO:
        codigo_terceiro, obs = 0, f"FPAS {fpas}:eiro, obs = None, "Simples Nacional — campo terceiros terceiros sempre zero"
    else:
        codigo_terceiro = terceiros_base
        removidos = []
        for ent in convenios:
            bit = ENTIDADES.get(ent.upper(),{}).get("codigo",0)
            if bit and (codigo_terceiro & bit):
                codigo_terceiro &= ~bit
                removidos.append(ent.upper())
        obs = f"Convênio: {', '.join(removidos)} removido(s)" if removidos else "Classificação padrão"
    rat_ajustado = round(rat_base * fap, 2)
    entidades    = decodificar_terceiros(codigo_terceiro) if codigo_terceiro else []
    total_terc   = sum(e["aliquota"] for e in entidades)
    return {
        "cnae":cnae_fmt,"fpas":fpas,"fpas_descricao":descricao,
        "codigo_terceiro":codigo_terceiro,"perc_acid_trabalho":rat_base,
        "codigo_sat":rat em branco"
    elif fpas in FPAS_SEMPRE_ZERO:
        codigo_terceiro, obs = 0, f"FPAS {fpas}: terceiros sempre zero"
    else:
        codigo_terceiro = terceiros_base
        removidos = []
        for ent in convenios:
            bit = ENTIDADES.get(ent.upper(),{}).get("_ajustado,"codigo_gps":cod_gps,"codigo_gfip":cod_gfip,
        "entidades":entidades,"total_terceiros_pcodigo",0)
            if bit and (codigo_terceiro & bit):
                codigo_terceiro &= ~bit
                removidos.append(ent.upper())
        obs = f"Convênio: {', '.join(removidos)} removido(s)" if removidos else "Classificação padrão"
    rat_ajustado = round(rat_base * fap, 2)
    entidades    = decodificar_terceiros(codigo_terceiro) if codigo_terceiro else []
    total_terc   = sum(e["aliquota"] for e in entidades)
    return {
        "cnae":cnae_fmt,"fpas":fpas,"fpas_descricao":descricao,
        "codigo_terceiro":codigo_terceiro,"perc_acid_trabalct":total_terc,"observacao":obs,"erro":None,
    }

# ──────────────────────────────────────────────────ho":rat_base,
        "codigo_sat":rat_ajustado,"codigo_gps":cod_gps,"codigo_gfip":cod_gfip,────────────────────────────
# CNPJ
# ──────────────────────────────────────────────────────────────────────────────
APIS
        "entidades":entidades,"total_terceiros_pct":total_terc,"observacao":obs,"erro = [
    "https://brasilapi.com.br/api/cn":None,
    }

# ──────────────────────────────────────────────────────────────────────────────
# CNPJ
# ──────────pj/v1/{cnpj}",
    "https://receitaws.com.br/v1/cnpj/{cnpj}",
    "https://m────────────────────────────────────────────────────────────────────
APIS = [
    "httpsinhareceita.org/{cnpj}",
]
HEADERS = {"User-Agent":"Mozilla/5.0 (://brasilapi.com.br/api/cnpj/v1/{cnpj}",
    "https://receitawsclassificador-fpas/2.0)"}

def limpar_cnpj(cnpj: str) -> str:
    return.com.br/v1/cnpj/{cnpj}",
    "https://minhareceita.org/{ re.sub(r"\D","",str(cnpj))

def validar_cnpj(cnpj: str) -> bool:
    ccnpj}",
]
HEADERS = {"User-Agent":"Mozilla/5.0 (classificador-fpas/2 = limpar_cnpj(cnpj)
    if len(c)!=14 or len(set(c).0)"}

def limpar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D","",str(cnpj))

def valid)==1: return False
    def calc_dv(digits,ar_cnpj(cnpj: str) -> bool:
    c = limpar_cnpj(cnpj)
    if len(c)!=14 pesos):
        soma=sum(int(d)*p for d,p in zip(digits,pesos)); restoor len(set(c))==1: return False
    def calc_=soma%11
        return 0 if resto<2 else 11-resto
    returndv(digits,pesos):
        soma=sum(int(d)*p (int(c[12])==calc_dv(c[:12],[5 for d,p in zip(digits,pesos)); resto=soma%11
        return 0 if resto,4,3,2,9,8,7,6,5,4,3,2]) and
            int(c[13])==calc_dv(c[:13],[6<2 else 11-resto
    return (int(c[12])==,5,4,3,2,9,8,7,6,5,4,3,2]))

def normalizar_resposta(data: dict,calc_dv(c[:12],[5,4,3,2,9,8,7,6,5,4,3,2]) cnpj: str) -> dict:
    cnae_codigo=(
        data.get("cnae and
            int(c[13])==calc_dv(c[:13],[6,5,4,3,2,9,8,7,6,5,4,3,2]))

def normal_fiscal") or data.get("cnae_fiscal_codigo")
        or (izar_resposta(data: dict, cnpj: str) -> dict:
    cnae_data.get("atividade_principal",[{}])[0].get("code","")
            if isinstancecodigo=(
        data.get("cnae_fiscal") or data.get("cnae_(data.get("atividade_principal"),list) else "")
    )
    cnae_descricao=(
        data.get("cnae_fiscal_descricao")
        or (data.get("atividade_principal",[{}])[0].get("text","")
            if isinstance(data.get("atividade_principal"),list) else "")
    )
    cnaefiscal_codigo")
        or (data.get("atividade_principal",[{}])[0].get("code","")
            if isinstance(data.get("atividade_principal"),list) else "")
    _str=str(cnae_codigo).zfill(7) if cnae_codigo else "")
    cnae_descricao=(
        data.get("cnae_fiscal_descricao")
        or (data.get("atividade_principal",[{}])[0].get("text
    cnae_fmt=f"{cnae_str[:4]}-{cnae_str[4]}/{cnae_str[5:]}" if len(cnae_str)==","")
            if isinstance(data.get("atividade_principal"),list) else "")
    )
    cnae_str=str7 else cnae_str
    simples=(
        data.get("opcao_pelo_simples")
        or (data.get("sim(cnae_codigo).zfill(7) if cnae_codigo else ""
    cnae_fmt=f"{cnae_str[:4]}-{cnae_strples",{}).get("optante",False)
            if isinstance(data.get("simples"),dict) else False)
    [4]}/{cnae_str[5:]}" if len(cnae_str)==7 else cnae_str
    simples=(
        data.get("opc)
    return {
        "cnpj":cnpj,
        "razao_social":data.get("razao_social") or data.get("nome","ao_pelo_simples")
        or (data.get("simples",{}).get("optante"),
        "nome_fantasia":data.get("nome_fantasia",""),
        "situacao":data.get("descricao_situacao_",False)
            if isinstance(data.get("simples"),dict) else False)
    )
    return {
        "cnpj":cnpj,
        "razcadastral") or data.get("situacao",""),
        "cnae_codigo":cnae_fmt,"ao_social":data.get("razao_social") or data.get("nome",""),
        "nome_fantasia":data.get("nome_fantasia",""),
        "situcnae_descricao":cnae_descricao,
        "simples":bool(simples),
        "natureza_juridica":data.get("descacao":data.get("descricao_situacao_cadastral") or data.get("situacao","ricao_natureza_juridica") or data.get("natureza_juridica",""),
        "porte":data.get("descricao_porte") or data.get("porte",""),"),
        "cnae_codigo":cnae_fmt,"cnae_descricao":cnae_descricao,
        "simples":bool
        "logradouro":data.get("logradouro",""),"numero":data.get("numero",""),
        "b(simples),
        "natureza_juridica":data.get("descricao_natureza_juridica") or data.get("natureza_juridica",""),
        "airro":data.get("bairro",""),"municipio":data.get("municipio","") or "",
        "uf":data.get("uf",""),"cep":data.getporte":data.get("descricao_porte") or data.get("porte",""),
        "logradouro":data.get("logradouro","("cep",""),
        "data_inicio":data.get("data_inicio_atividade") or data.get("abertura",""),"numero":data.get("numero",""),
        "bairro":data.get("bairro",""),"municipio":data.get("municipio","")"),
        "erro":None,
    }

def consultar_cnpj(cnpj_raw or "",
        "uf":data.get("uf",""),"cep":data.get("cep",""),
        "data_inicio":data.get("data_inicio_: str, delay: float=1.0) -> dict:
    cnpj=limpar_cnpj(cnpj_raw)
    if not validatividade") or data.get("abertura",""),
        "erro":None,
    }

def consular_cnpj(cnpj): return {"erro":f"CNPJ inválido: {cnpj_raw}"}
    timetar_cnpj(cnpj_raw: str, delay: float=1.0) -> dict:
    cnpj=.sleep(delay)
    for url_tpl in APIS:
        try:
            r=limpar_cnpj(cnpj_raw)
    if not validar_cnpj(cnpj): return {"erro":f"CNPJ inválrequests.get(url_tpl.format(cnpj=cnpj),headers=HEADERS,timeout=12)
            if r.status_code==200: return normalido: {cnpj_raw}"}
    time.sleep(delay)
    for url_tpl in APIS:
        izar_resposta(r.json(),cnpj)
            if r.status_code==429: time.sleep(60)
        try:
            r=requests.get(url_tpl.format(cnpj=cnpj),headers=HEADERS,timeout=12)
            if r.statusexcept Exception: continue
    return {"erro":f"CNPJ {cnpj} não encontrado em n_code==200: return normalizar_resposta(r.json(),cnpj)
            if r.status_code==429: timeenhuma API."}

def extrair_cnpjs_do_texto(texto: str).sleep(60)
        except Exception: continue
    return {"erro":f"CNPJ {cnpj} -> list[str]:
    encontrados=re.findall(r"\d{2}[\ não encontrado em nenhuma API."}

def extrair_cnpjs.\s]?\d{3}[\.\s]?\d{3}[\/\s]?\d{4_do_texto(texto: str) -> list[str]:
    encontrados=re.findall(r"\d}[-\s]?\d{2}",texto)
    vistos,unicos=set(),[]
    {2}[\.\s]?\d{3}[\.\s]?\d{3}[\/for c in encontrados:
        limpo=limpar_cnpj(c)
        if limpo not in vistos:
            vistos.add(limpo); unicos.append\s]?\d{4}[-\s]?\d{2}",texto)
    vistos,(c.strip())
    return unicos

# ──────────────────────────────────────────────────────────────────────────────unicos=set(),[]
    for c in encontrados:
        limpo=limpar_cnpj(c)
        if limpo not in vistos:
            vis
# HELPERS DE DATA — todas as datas no formato DDtos.add(limpo); unicos.append(c.strip())
    return unicos

# ──────────────────────────────────────────────────/MM/AAAA
# ──────────────────────────────────────────────────────────────────────────────
def _formatar_data(data_raw: str────────────────────────────
# LEIAUTE FOSERVICOS —) -> str:
    """Converte qualquer data para DD/MM/AAAA."""
    from datetime import datetime
    s 25 colunas
# ──────────────────────────────────────────────────────────────────────────────
C = str(data_raw or "01/01/2020")[:10]
    forOLUNAS_LEIAUTE = [
    "Codigo_empresa","Codigo_Servicos fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return "01/01/2020"

def _data_para","CNPJ_CPF","Tipo_Inscricao",
    "Codigo_Terc_vigencia(d: date) -> str:
    """date → DD/MM/AAAA eiro","Perc_Acidente_Trabalho","Codigo_FPAS","CNAE",
    "Codigo_GFIP","Codigo_GPS","Nome","Endereco","Numero00:00:00  (formato FOVIGENCIAS","Bairro",
    "CEP","Cidade","Estado","Codigo_Filial","Sequ)"""
    return d.strftime("%d/%m/%Y") + " 00:00:00"

# Data fiencia_GPS","Tipo",
    "Codigo_Municipio","Data_Inicio","Situxa de competência fim: 31/12/3000
COMP_FIM = "acao","Codigo_eSocial","Origem_Reg",
]31/12/3000 00:00:00"

# ──────────────────────────────────────────────────────────────────────────────
# LEIAUTE

def _formatar_data(data_raw: str) -> str:
    from datetime import datetime
    s FOSERVICOS — 25 c=str(data_raw or "2020-01-01")[:10]
    for fmt in ("%dolunas
# ──────────────────────────────────────────────────────────────────────────────
COLUNAS_LEIAUTE = [
    "Codigo_/%m/%Y","%Y-%m-%d","%d-%m-%Y"):
        try: return datetime.strptime(sempresa","Codigo_Servicos","CNPJ_CPF","Tipo,fmt).strftime("%Y-%m-%d")
        except ValueError: continue
    return "2020-01-01"

def mon_Inscricao",
    "Codigo_Terceiro","Perc_Acidente_tar_linha_dominio(r: dict, tipo_Trabalho","Codigo_FPAS","CNAE",
    "Codigo_GFIP","Codigo_GPScod: int, cod_servico: int, codigo_empresa: int=","Nome","Endereco","Numero","Bairro",
    "CEP","Cidade","Estado1) -> list:
    cnpj_limpo=limpar_cnpj(r.","Codigo_Filial","Sequencia_GPS","Tipo",get("cnpj",""))
    cod_terceiro=r.get("codigo_terceiro","")
    if isinstance(cod_terceiro,int
    "Codigo_Municipio","Data_Inicio","Situacao","Codigo_): cod_terceiro=f"{cod_terceiro:04d}"
    elif cod_terceiro is None: cod_terceiro=""
    ceSocial","Origem_Reg",
]

def montar_linhaep=re.sub(r"\D","",str(r.get("cep","") or ""))
    municipio=str_dominio(r: dict, tipo_cod: int(r.get("municipio","") or "")
    uf=str(r.get("uf","") or "")
    cod_mun=buscar_codigo_municipio(, cod_servico: int, codigo_empresa: int=1municipio,uf)
    data_ini=_formatar_data(r.get("data_inicio",""))
    return [
        codigo) -> list:
    cnpj_limpo=limpar_cnpj(r.get("cn_empresa,cod_servico,cnpj_limpo,1,
        cod_terceiro,
        str(r.get("perc_acidpj",""))
    cod_terceiro=r.get("codigo_terceiro","")
    if isinstance(cod_terceiro,int):_trabalho","") or ""),
        str(r.get("codigo_fpas","") or ""),
        str(r.get("cnae_codigo","") or ""),
         cod_terceiro=f"{cod_terceiro:04d}"
    elif cod_terceiro is None: cod_terceiro=""
    cepstr(r.get("codigo_gfip","") or ""),
        str(r.get("codigo_gps","") or ""),
        str(r.get("razao_social","") or ""),
        str(r.=re.sub(r"\D","",str(r.get("cep","") or ""))
    municipio=str(r.get("municipioget("logradouro","") or ""),
        str(r.get("numero","") or ""),
        str(r.get("bairro","") or ""),
        cep,municip","") or "")
    uf=str(r.get("uf","") or "")
    cod_mun=buscar_codigo_municipio(municipio,io,uf,
        cod_servico,1,tipo_cod,cod_mun,data_ini,1,coduf)
    data_ini=_formatar_data(r.get("data_inicio",""))   # DD/MM/AAAA
    return_servico,"",
    ]

def gerar_txt_leiaute(linhas: list[list]) [
        codigo_empresa,cod_servico,cnpj_limpo,1,
        cod_terceiro,
        str(r -> bytes:
    linhas_txt=[]
    for campos in linhas:
        row=[str(v.get("perc_acid_trabalho","") or ""),
        str(r.get("codigo_fpas","") or ""),
        str(r.get("cnae_codigo","") or ""),
        str(r.get("codigo_gfip","") or ""),
        str(r.get("codigo_gps","") or ""),
        str(r.get("razao_social","") or ""),
        str(r.get("logradouro","") or ""),
        str(r.) if v is not None else "" for v in campos]
        while len(row)<25: row.append("")
        linhas_txt.append("\t".join(row[:25]))
    return ("\r\n".join(linhas_txt)+"\r\n").encode("utf-8")

# ──────────────────────────────────────────────────────────────────────────────
# LEIAUTE FOVIGENCIAS_get("numero","") or ""),
        str(r.get("bairro","") or ""),
        cep,municipio,uf,
        cod_servSERVICOS — fiel ao xlsx real
# Campos corico,1,tipo_cod,cod_mun,data_ini,1,cod_servico,"rigidos vs versão anterior:
#   i_fil",
    ]

def gerar_txt_leiaute(linhas: list[list]) -> bytes:
    liniais   = 1           (no xlsx real éhas_txt=[]
    for campos in linhas:
        row=[str(v) if v is not None else "" for sempre 1, não cod_servico)
#   origem_reg  = 0 v in campos]
        while len(row)<25: row.append("")
        linhas_txt.append("\t".join(row[:25]))
    return ("\r           (no xlsx real é 0=Sistema, não 2)
#   I_TERC\n".join(linhas_txt)+"\r\n").encode("utf-8")

# ──────────────────────────────────────────────────────────────────────────────
# LEIAUTEEIROS = codigo_terceiro (valor numérico inteiro da tabela, FOVIGENCIAS_SERVICOS — ordem exata do xlsx ex: 3, 115, 162)
#   I_CN real
# Correções vs. versão anteriorAE_ESOCIAL = mesmo valor de I_CNAE20:
#   • origem_reg = 0  (todos (quando preenchido)
#   TIPO_ENDERECO  = 0         os registros reais têm 0, não 2)
#   • i_filiais(padrão seguro para importação)
#   TIPO_LOTACAO_  = 1  (todos os registros reais têm 1, não cod_servico)
#   • VIGENCIA eESOCIAL = 1  (padrão para empresas com eS COMPETENCIA_FIM_VIGENCIA em DD/MM/ocial ativo)
# ──────────────────────────────────────────────────────────────────────────────
def montar_linha_vigencia(r: dict, codAAAA HH:MM:SS
# ──────────────────────────────────────────────────────────────────────────────
def montar_linha_vigencia(r: dict, cod_servico: int, tipo_cod: int,
                          codigo_empresa: int, vigencia: str,
                          desc_servico: int, tipo_cod: int,
                          codigo_empresa: int, vigencia: str,
                          descricao_vigencia: strricao_vigencia: str, cod_mun: str) -> list:

    cnpj_limpo =, cod_mun: str) -> list:
    cnpj_limpo = limpar_cnpj(r.get("cnpj limpar_cnpj(r.get("cnpj",""))

    # codigo_terceiro numérico (ex: 3, 79, 115)",""))

    cod_terc = r.get("codigo_terceiro","")
    if isinstance(cod_terc,int):    → string
    cod_terc_raw = r.get("codigo_terceiro","")
    if isinstance(cod_terc_str = str(cod_terc)
    elif cod_terc is None:         cod_terc_str = "0cod_terc_raw, int):
        cod_terc_str = str(cod_terc_raw)
        i"
    else: cod_terc_str = str(cod_terc).strip() if str(cod_terc).strip_terceiros  = cod_terc_raw          # I_TERCEIROS = valor numérico
    elif cod() else "0"

    entidades  = r.get("entidades",[]) or []
    perc_te_terc_raw is None:
        cod_terc_str = "0"
        i_terceiros  = 0
    else:
        s = str(cod_tercrc  = sum(e.get("aliquota",0) for e in entidades)

    rat   _raw).strip()
        cod_terc_str = s if s else "0"
        try: i_terceiros = int(s)= r.get("perc_acid_trabalho",0) or 0
    fpas  = r.get("codigo_fpas", if s else 0
        except: i_terceiros = 0

    # % terceiros = soma das0) or 0
    gfip  = r.get("codigo_gfip","") or "115"
    gps   = r.get("codigo_g alíquotas das entidades decodificadas
    entidades  = r.get("entidadesps","")  or "2100"
    end   = str(r.get("log", []) or []
    perc_terc  = sum(e.get("aliquota", 0) for e in entradouro","") or "")
    num   = str(r.get("numero","")   or "0")
    bairro= str(r.get("bidades)

    rat        = r.get("perc_acid_trabalho", 0) or 0
    fpas       = r.get("codigo_airro","")   or "")
    cep   = re.sub(r"\D","",str(r.get("cep","") or ""))
    mun   = str(r.fpas", 0) or 0
    gfip       = r.get("codigo_gfip", "") or "115"
    gps        =get("municipio","") or "")
    uf    = str(r.get("uf","")        or "")

    cnae_fmt r.get("codigo_gps",  "") or "2100"
    endereco   = str(r.get("logradouro"," = str(r.get("cnae_codigo","") or "")
    cnae_num = re.sub(r"\D","",cnae_fmt)   ") or "")
    numero     = str(r.get("numero","") or "0")
    bairro     = str(r.get("bairro","") or "")
    cep        = re# ex: "4761003"

    # TIPO_.sub(r"\D","",str(r.get("cep","") or ""))
    municipio  = str(r.get("municipio","") or "")
    uf         = str(r.get("uf","")INFORMACAO_ALIQUOTA_ACIDENTE_TRABALHO: 1=Inform or "")

    # CNAE numérico (7 dígitos sem formatação)ado (padrão importação)
    tipo_inf para I_CNAE20 e I_CNAE_ESOCIAL
    cnae_fmt   = str_aliq = 1

    return [
        # ──(r.get("cnae_codigo","") or "")
    cnae_num   = re.sub(r"\D","",cnae_fmt)   # ex: " Identificação ──────────────────────────────────────────────────
        codigo_empresa,          # codi_emp4761003"

    comp_fim = "
        cod_servico,             # i_servicos
        vigencia,                # VIGENCIA          DD/MM/AAAA3000-12-31 00:00:00"

    return 00:00:00
        descricao_vigencia,      # DESCRICAO
        cnpj_limpo,              # cgc [
        # ── Identificação ─────────────────────────────────────────────────
        codigo_empresa,          
        1,                       # tipo_insc         1=CN# codi_emp
        cod_servico,             # i_servicos
        vigPJ
        # ── Alíquotas ─────────────────────────────────────────────────────
        codencia,                # VIGENCIA          ex: "2020-01-01 00:00:00"
        descricao_vigencia,      #_terc_str,            # codigo_terceiro
        perc_terc,               # perc_terceiro
        0,                       # perc_in DESCRICAO
        cnpj_limpo,              # cgc
        1,                       # tipo_inss_empresa
        rat,                     # perc_acid_trabalho
        0,                       # codigo_sat
        0,                       # perc_autonomos
        fpas,                    # codigo_fpas
        0,                       # codigo_atividade
        gfip,                    # codigo_gfip
        gps,                     # codigo_gps
        # ── Bancosc         1=CNPJ
        # ── Alíquotas ────────────────────────────────────────────────────
        cod_terc_str,            # codigo_terceiro    (valor numérico em string)
        perc_terc,               # perc_terceiro      (soma das alíquotas)
        0,                       # pe / FGTS ──────────────────────────────────────────────────
        0,                       # i_banrc_inss_empresa
        rat,                     # perc_acid_trabalho
        0,                       # codigocos
        0,                       # numero_fgts
        # ── Endereço ──────────────────────────────────────────────────────
        end,                     #_sat
        0,                       # perc_autonomos
        fpas,                    # codigo_fpas
        0,                       # codigo_atividade   (sempre 0 no endereco
        num,                     # numero
        bairro,                  # bairro
        cep,                     # cep
        mun,                     # cidade
        uf,                      # estado
        # ── Filial / GPS ──────────────────────────────────────────────────
         xlsx real)
        gfip,                    # codigo_gfip
        gps,                     # codigo_gps
        # ── Banco / FGTS ─────────────────────────────────────────────────
        0,                       # i_bancos
        0,                       1,                       # i_filiais  ← sempre 1 (conforme xlsx real)
        0,                       # sequ# numero_fgts
        # ── Endereço ─────────────────────────────────────────────────────
        endereco,                # enderencia_gps
        0,                       # filantropia
        0,                       # origem_reg ← sempre 0 (conforme xlsx real)
        tipo_cod,                # tipo
        # ── Município ─────────────────────────────eco
        numero,                  # numero
        bairro,                  # bairro
        cep,                     # cep
        municipio,               # cidade
        uf,                      ────────────────────────
        0,                       # codi_mun
        cod_mun,                 # codigo_municipio
        # ──# estado
        # ── Filial / GPS ─────────────────────────────────────────────────
         FPAS / Lei 12.546 ─1,                       # i_filiais          ← CORRIGIDO: sempre 1 (xlsx real)
        0,                ────────────────────────────────────────────
        0,                       # I_FPAS
        0,                       # CALCULA_IN       # sequencia_gps
        0,                       # filantropia
        0,                       # origem_reg         ← CORRIGIDO:SS_EMPRESA_LEI_12546
        # ── Tomador / End 0=Sistema (xlsx real)
        tipo_cod,                # tipo
        # ── Município ────ereço ────────────────────────────────────────────
        0,                       # TIPO_SERVICO_TOMADOR
        0────────────────────────────────────────────────
        0,                       # codi_mun           (sempre 0 no xlsx real),                       # TIPO_ENDERECO
        "",                      # COMPLEMENTO
        # ── Empr
        cod_mun,                 # codigo_municipio
        # ── FPAS / Lei eitada parcial ────────────────────────────────────────────
        0,                       # TIPO12.546 ────────────────────────────────────────────
        0,                       # I_FPAS_INSCRICAO_CONTRATANTE_EMPREITADA_PARCIAL
        "",                      # INSCRICAO_CONTRATANTE_EMPREITADA_
        0,                       # CALCULA_INSS_EMPRESA_LEPARCIAL
        "",                      # NOME_CONTRATANTE_EMPREITADA_PARCIAL
        0,                       # TIPO_INSCRICAO_PROPRIETI_12546
        # ── Tomador / Endereço ───────────────────────────────ARIO_CEI_EMPREITADA_PARCIAL
        "",                      # INSCRICAO_PROPRIETARIO_CEI_EMPREITADA_PARCIAL
        "",                      ────────────
        0,                       # TIPO_SERVICO_TOMADOR
        0,                       # TIPO_ENDERECO      ← CORRIGIDO: # NOME_PROPRIETARIO_CEI_EMPREITADA_PARCIAL
        # ── CNAE /0 (padrão seguro)
        "",                      # COMPLEMENTO
        # ── Em RAT ────────────────────────────────────────────────────
        cnae_num,                # I_CNAE20
        tipopreitada parcial ───────────────────────────────────────────
        0,_inf_aliq,           # TIPO_INFORMACAO_ALIQUOTA_ACIDENTE_TRABALHO
        0,                       # I_PROCESSO
        0,                       # I                       # TIPO_INSCRICAO_CONTRATANTE_EMPREITADA_PARCIAL
        "",                      # INSCRICAO_CONTRATANTE_SCP
        "",                      # DDD
        "",                      # TELEFONE
        COMP_FIM,                #_EMPREITADA_PARCIAL
        "",                      # NOME_CONTRATANTE_EMPREITADA_PARCIAL
        0,                       # TIPO_INSCRICAO_PROPRIET COMPETENCIA_FIM_VIGENCIA  DD/MM/AAAA 00:00:00
        0,                       # TIPO_LOTARIO_CEI_EMPREITADA_PARCIAL
        "",                      # INSCRICAO_PROPRIETARIO_CEI_EMPREITADA_PARCIAL
        "",                      # NOME_PROPRIETARIO_CEI_EMPREITADA_PARCIAL
        # ── CNAE / RAT ───ACAO_ESOCIAL  ← 0 (padrão importação)
        0,                       # I_PROCESSO_TERCEIROS
        "",                      # CAEPF────────────────────────────────────────────────
        cnae_num,                # I_CNAE20           (numérico sem formatação)
        1,                       # TIPO_INFORMACAO_ALIQUOTA_ACIDENTE_TRABALHO (1=Inform
        0,                       # TIPO_CAEPF
        0,                       # REGISTRO_PONTO
        1,                       # CONTRATACAO_APRENDIZ  1=Dispensado)
        0,                       # I_PROCESSO
        0,                       # I_SCP
        "",ado
        0,                       # I_PROCESSO_CONTRATACAO_APRENDIZ
        0,                       # REALIZA_CONTRATACAO_                      # DDD
        "",                      # TELEFONE
        comp_fim,                # COMPETENCIA_FAPRENDIZ_INTERMEDIO_...
        "",                      # CODIGO_SUSPIM_VIGENCIA  (3000-12-31 00:00:00)
        1,                       # TIPO_ENSAO_PROCESSO_RAT
        0,                       # SOMA_CODIGOS_SUSPENSAO_TERCLOTACAO_ESOCIAL  ← 1 para eSocial ativo
        0,                       # IEIROS
        perc_terc,               # PERCENTUAL_TERCEIRO_BRUTO
        #_PROCESSO_TERCEIROS
        "",                      # CAEPF
        0 ── S-1005 ────────────────────────────────────────────────────────
        0,                       #,                       # TIPO_CAEPF
        0,                       # REGISTRO_PONTO
        1,                       # CONT I_DADOS_EVENTOS_ESOCIAL_S_1005   **SEM USO**
        0,                       # I_LOTE_ESOCIAL_S_1005            **SEM USO**RATACAO_APRENDIZ  (1=Dispensado)
        0,                       # I_PROCESSO_CONTRATACAO_APRENDIZ
        0,                       # REALIZA
        0,                       # STATUS_ESOCIAL_S_1005            **SEM USO**
        1,                       # ENVIAR_CONTRATACAO_APRENDIZ_INTERMEDIO_...
        "",                      # CODIGO_ESOCIAL_S_1005
        0,                       # INCLUSAO_VALIDADA_ESOCIAL_S_1_SUSPENSAO_PROCESSO_RAT
        0,                       # SOMA_CODIGOS_SUSPENSAO_TERC005 **SEM USO**
        0,                       # GERAR_RETIFICACAO_ESOCIAL_S_1005EIROS
        perc_terc,               # PERCENTUAL_TERCEIRO_BRUTO
        # ── S
        # ── S-1020 ────────────────────────────────────────────────────────
        0,                       # I_DADOS_EVENTOS_ESOCIAL-1005 ───────────────────────────────────────────────────────
        0,                       # I_S_1020   **SEM USO**
        0,                       # I_LOTE_ESOCIAL_S_1020            **SEM USO**
        0,                       # STATUS_ESOCIAL_S_DADOS_EVENTOS_ESOCIAL_S_1005   **SEM USO**
        0,                       # I_LOTE_ESOCIAL_S_1005            _1020            **SEM USO**
        1,                       # ENVIAR_ESOCIAL_S_1020
        0,                       # INCLUSAO_VALIDADA_ESOCIAL_S_1020 ****SEM USO**
        0,                       # STATUS_ESOCIAL_S_1005            **SEM USO**
        1,                       # ENVIARSEM USO**
        0,                       # GERAR_RETIFICACAO_ESOCIAL_S_1020
        # ── eSocial extra_ESOCIAL_S_1005
        0,                       # INCLUSAO_VALIDADA_ESOCIAL_S_1 ─────────────────────────────────────────────────
        cnae_num,                # I_CNAE_ESOCIAL
        0,                       # I005 **SEM USO**
        0,                       # GERAR_RETIFICACAO_ESOCIAL_S_1005_TERCEIROS
        0,                       # PROCESSAR_EXCLUSAO_ESOCIAL_S
        # ── S-1020 ───────────────────────────────────────────────────────
        0,                       # I_DADOS_EVENTOS_ESOCIAL_1005
        0,                       # PROCESSAR_EXCLUSAO_ESOCIAL_S_1020
        1,                       # ORIGEM_ALTERACAO
        # ── S-1080 ────────────────────────────────────────────────────────
        0,                       # I_DADOS_EVENTOS_ESOCIAL_S_1020   **SEM USO**
        0,                       # I_LOTE_ESOCIAL_S_1020            **SEM USO**
        0,                       # STATUS_ESOCIAL_S_1020            **SEM USO**
        1,                       # ENVIAR_ESOCIAL_S_1020
        0,                       # INCLUSAO_VALIDADA_ESOCIAL_S_1020 **_S_1080   **SEM USO**
        0,                       # I_LOTE_ESOCIAL_S_1080            **SEM USO**
        0,                       # STATUS_ESOCIAL_SSEM USO**
        0,                       # GERAR_RETIFICACAO_ESOCIAL_S_1020
        # ── eSocial extra ────_1080            **SEM USO**
        1,                       # ENVIAR_ESOCIAL_S_1080
        0,                       # INCLUSAO_VALIDADA_ESOCIAL_S_1080 **────────────────────────────────────────────
        cnae_num,                # I_CNAE_ESOCIAL     ← CORRIGIDO: igual I_CNAE20
        SEM USO**
        0,                       # GERAR_RETIFICACAO_ESOCIAL_S_1080
        0,                       # PROCESSAR_EXCLUSAO_ESOCIAL_S_1080i_terceiros,             # I_TERCEIROS        ← CORRIGIDO: valor numérico inteiro
        0,                       #
        # ── Retenção / COMPANY_ID ───────────────────── PROCESSAR_EXCLUSAO_ESOCIAL_S_1005
        0,                       # PROCESSAR_EXCLUSAO────────────────────
        0,                       # EFETUAR_RETENCAO_INSS_NOTAS_FISC_ESOCIAL_S_1020
        1,                       # ORIGEM_ALTERACAO   (1=Sistema)AIS_INSCRICAO_OUTROS_CLIENTES
        "{00000000-0000-0000-0000-000
        # ── S-1080 ───────────────────────────────────────────────────────
        0,                       # I_DADOS_EVENTOS000000000}",  # COMPANY_ID
        "",                      # NUMERO_RECIBO_ESOCIAL_VALID_ESOCIAL_S_1080   **SEM USO**
        0,                       # I_LOTE_ESOCIAL_S_1080            **SEM USO**
        0,                       # STATUS_ACAO_AUTOMATICA_1005
        "",                      # NUMERO_PROCESSO_APRENDIZ
        0,                       # INEXIGIBILIDADE_RAT
        0,                       #ESOCIAL_S_1080            **SEM USO**
        1,                       # ENVIAR_ESOCIAL_S_1080
        0,                       # INCLUSAO_VALIDADA_ESOCIAL_S CALCULAR_APOIO_FINANCEIRO_FOLHA_COLABORADOR__1080 **SEM USO**
        0,                       # GERAR_RETIFICACAO_ESOCIAL_S_1080
        0,                       # PROCESSAR_EXCLUSAO_ESRS_MTE_991_2024OCIAL_S_1080
        # ── Final ────────────────────────────────────────────────────────
        0,                       # EF
    ]

def _linha_vigencia_vaziaETUAR_RETENCAO_INSS_NOTAS_FISCAIS_INSCRICAO_OUTROS(cnpj: str, cod_servico: int, tipo_cod: int,
                          codigo_empresa: int, vigencia: str,
                          descricao_vigencia:_CLIENTES
        "{00000000-0000-0000-0000-000000000000}",  # COMPANY str) -> list:
    return montar_linha_vigencia(
        r={"cnpj":cnpj,"codigo_ID
        "",                      # NUMERO_RECIBO_ESOCIAL_VALID_terceiro":0,"perc_acid_trabalho":0,
           "codigo_fpas":0,"codigo_gfip":"115","codigo_gps":"ACAO_AUTOMATICA_1005
        "",                      # NUMERO_PROCESSO2100",
           "logradouro":"","numero":"0","bairro":"","cep":"",
           "municipio":"","uf":"","cnae_codigo_APRENDIZ
        0,                       # INEXIGIBILIDADE_RAT
        0,                       # CALCULAR_APOIO_FINANCEIRO_FOLHA_COLABORADOR_":"","entidades":[]},
        cod_servico=cod_servico, tipo_cod=tipo_cod,
        codigo_empresa=codigo_empresa,RS_MTE_991_2024
        vigencia=vigencia, descricao_vigencia=descricao_vigencia,
        cod_mun="",
    )

def gerar_txt_vig
    ]

def _linha_vigencia_vazia(cnpencias(linhas: list[list]) -> bytes:
    linhas_txt=[]
    for campos in linhas:
        row=[str(v) if v is not None else "" forj: str, cod_servico: int, tipo_cod: int,
                          codigo_empresa: int, vigencia: str,
                          descricao_vigencia: str) -> list: v in campos]
        linhas_txt.append("\t".join(row))
    return ("\r\n".join(linhas_txt)+"\r\n").encode("utf-8")

def gerar_excel_conferencia(df: pd.DataFrame, df_err: pd.DataFrame | None) -> bytes:
    buf
    return montar_linha_vigencia(
        r={"cnpj":cnpj,"=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as writer:
        df.to_excel(writer,sheet_name="Importcodigo_terceiro":0,"perc_acid_trabalho":0,
           "codigo_fpas":0,"codigo_gfip":"acao_Dominio",index=False)
        if df_err is not None and len(df_err)>0:
            df_115","codigo_gps":"2100",
           "logradouro":"","numero":"0","bairro":"","cep":"",
           "municipio":"","uf":"","cnae_codigoerr.to_excel(writer,sheet_name="Erros",index=False)
        wb=writer.book; ws=wb["":"","entidades":[]},
        cod_servico=cod_servico,tipo_cod=tipo_cod,
        codigo_empresa=codigo_empresa,vigencia=vigencia,
        descricao_vigencia=descricao_vigencia,cod_mun="",
    )

def gerar_txt_vigencias(linhas: list[listImportacao_Dominio"]
        from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
        from openpyxl.utils import get_column_letter
        HDR_FILL=PatternFill("solid",fgColor="FF8]) -> bytes:
    linhas_txt=[]
    for campos in linhas:
        row=[str(v) if v is not None else "" for v in campos]
        linhas_txt.append("\t000")
        OK_FILL=PatternFill("solid",fgColor="0D2B0".join(row))
    return ("\r\n".join(linhas_txt)+"\r\n").encode("utf-8")

def gerar_excel_conferD")
        ERR_FILL=PatternFill("solid",fgColor="2B0D0D")
        ODDencia(df: pd.DataFrame, df_err: pd.DataFrame | None) -> bytes:
    buf=io.BytesIO()
    with pd.Excel_FILL=PatternFill("solid",fgColor="242424")
        EVEN_FILL=PatternFill("solid",fgColor="2EWriter(buf,engine="openpyxl") as writer:
        df.to_excel(writer,sheet_name="Importacao_Dominio",index=False)2E2E")
        HDR_FONT=Font(bold=True,color="FFFFFF",size=9)
        CELL
        if df_err is not None and len(df_err)>0:
            df_err.to_excel(writer,sheet_name="Erros",index=False)_FONT=Font(color="F0F0F0",size=9)
        THIN=Side(style="thin",
        wb=writer.book; ws=wb["Importacao_Dominio"]
        fromcolor="3A3A3A")
        BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
         openpyxl.styles import PatternFill,Font,Alignment,Border,Side
        from openpyxl.utils import get_column_letter
        HDfor col_idx,cell in enumerate(ws[1],start=1):
            cell.R_FILL=PatternFill("solid",fgColor="FF8000")
        OK_FILL=PatternFfill=HDR_FILL; cell.font=HDR_FONT
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=BORDERill("solid",fgColor="0D2B0D")
        ERR_FILL=PatternFill("solid",fg
            ws.column_dimensions[get_column_letter(col_idx)].width=18
        ws.row_dimensions[1].height=36Color="2B0D0D")
        ODD_FILL=PatternFill("solid",fgColor="242
        try: status_col=df.columns.tolist().index("_424")
        EVEN_FILL=PatternFill("solid",fgColor="2E2E2E")
        HDR_FONT=Font(status")+1
        except ValueError: status_col=None
        for row_idx,row in enumerate(ws.iter_rows(min_row=2),bold=True,color="FFFFFF",size=9)
        CELL_FONT=Font(color="F0F0F0",sizestart=2):
            sv=str(ws.cell(row=row_idx,column=status_col).value or "") if status=9)
        THIN=Side(style="thin",color="3A3A3A")
        BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
        for col_idx_col else ""
            bg=OK_FILL if sv=="OK" else (ERR_FILL if "ERRO" in sv else (EVEN_FILL if row_idx%2==0 ,cell in enumerate(ws[1],start=1):
            cell.fill=HDR_FILL; cell.font=HDR_else ODD_FILL))
            for cell in row:
                cell.fill=bg; cell.font=CELL_FONT
                cell.alignment=Alignment(vertical="center");FONT
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=BORDER
            ws.column_dimensions[get_column_letter(col_idx cell.border=BORDER
        ws.freeze_panes="A2"
    return buf.getvalue()

# ──────────────────────────────────────────────────)].width=18
        ws.row_dimensions[1].height=36
        try:────────────────────────────
# HELPERS UI
# ──────────────────────────────────────────────────────────────────────────────
def metric status_col=df.columns.tolist().index("_status")+1
        except ValueError_card(value,label,cls=""):
    return f'<div class="tr-metric {: status_col=None
        for row_idx,row in enumerate(ws.iter_rows(min_row=2),start=2):
            svcls}"><div class="tr-metric-value">{value}</div><div class="tr-metric-label">{label}</div></div>'

def result_item(label=str(ws.cell(row=row_idx,column=status_col).value or "") if status_col else ""
            bg,value,highlight=False):
    cls="highlight" if highlight else ""
    return f'<div class="result-item"><div class="result-item=OK_FILL if sv=="OK" else (ERR_FILL if "ERRO" in sv else (EVEN_FILL if row_idx%2==0 else ODD_FILL))-label">{label}</div><div class="result-item-value {cls}">{value}</div></div>'

# ──────────────────────────────────────────────────────────────────────────────
# APP
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Classific
            for cell in row:
                cell.fill=bg; cell.font=CELL_FONT
                cell.alignment=Alignment(vertical="center"); cell.border=BORDER
        ws.freeze_panes="A2"
    return buf.getvalue()

# ──────────────────────────────────────────────────────────────────────────────ador FPAS | Domínio Sistemas",
    page_icon="🏛
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────
def metric_card(value️", layout="wide", initial_sidebar_state="expanded"
)
st.markdown(CSS, unsafe_allow_html=True)

st,label,cls=""):
    return f'<div class="tr-metric {cls}.markdown("""
<div class="tr-header">
    <div class="tr-logo-box">TR"><div class="tr-metric-value">{value}</div><div class="tr-metric-label">{label}</div></div>'

def result_item(label,value,highlight</div>
    <div>
        <div class="tr-title">Classificador FPAS / Terceiros=False):
    cls="highlight" if highlight else ""
    return f'<div class="result-item"><div class="result-item-label">{label}</div / SEFIP</div>
        <div class="tr-subtitle">DOMÍNIO SISTEMAS &nbsp><div class="result-item-value {cls}">{value}</div></div>'

# ──────────────────────────────────────────────────────────────────────────────
# APP;·&nbsp; Thomson Reuters &nbsp;·&nbsp; IN RFB nº 971
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Classificador/2009</div>
    </div>
    <div class="tr-badge">v8.2 FPAS | Domínio Sistemas",
    page_icon="🏛️", layout="wide",</div>
</div>
""", unsafe_allow_html=True)

# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown initial_sidebar_state="expanded"
)
st.markdown(CSS, unsafe_allow_html=True)

st.markdown("""
<div(f'<p style="color:{TR_ORANGE class="tr-header">
    <div class="tr-logo-box">TR</div>
    <div>};font-size:13px;font-weight:700;letter-spacing:1px;margin-bottom:16px;">
        <div class="tr-title">Classificador FPAS / Terceiros / SEFIP⚙ CONFIGURAÇÕES</p>', unsafe_allow_html=True)
    st.markdown(f'<p style="font-size:10px;font-weight:700</div>
        <div class="tr-subtitle">DOMÍNIO SISTEMAS &nbsp;·&nbsp; Thomson Reuters &;color:{TR_ORANGE};text-transform:uppercase;letter-spacing:1px;">🏢 Empresa no Domínio</p>',nbsp;·&nbsp; IN RFB nº 971/2009</div>
    </div unsafe_allow_html=True)
    codigo_empresa_dom = st.number_input(
        "Código Interno>
    <div class="tr-badge">v8.2</div>
</div>
""", unsafe_allow_html=True)

# da Empresa", min_value=1, max_value=999999,
        value=st.session_state.get(" ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(fcodigo_empresa_dom_val",1), step=1,
        key="codigo_empresa_dom_input",'<p style="color:{TR_ORANGE};font-size:13
        help="Código da empresa no Domínio Sistemas."
    )
    st.session_state["codigo_empresa_dom_valpx;font-weight:700;letter-spacing:1px;margin-bottom:16px;">⚙ CONFIGURAÇÕES</p>', unsafe_"] = codigo_empresa_dom

    st.divider()
    st.markdown(f'<p style="font-size:10px;font-weight:700;color:{TR_ORANGE};text-transform:uppercase;letter-allow_html=True)
    st.markdown(f'<p style="font-size:10px;font-weight:700;color:{TR_ORANGE};textspacing:1px;">📊 Parâmetros SEFIP</p>', unsafe_allow_html=True)
    f-transform:uppercase;letter-spacing:1px;">🏢 Empresa no Domínio</p>', unsafe_allow_html=Trueap   = st.number_input("FAP",           min_value=0.5, max_value=2.0, value=1.0, step=0.01)
    delay = st.number_input(")
    codigo_empresa_dom = st.number_input(
        "Código Interno da Empresa", min_value=1, max_value=999999,
        value=st.session_state.get("codigo_empresa_dom_val",Intervalo (s)", min_value=0.3, max_value=5.0, value=1.0, step=0.1)
    st.markdown(f'<p style="1), step=1,
        key="codigo_empresa_dom_input",
        help="Código da empresa no Domínio Sistemas."font-size:10px;font-weight:700;color:{TR_ORANGE};text-transform:uppercase;letter-spacing:1px;margin-top:12px;">
    )
    st.session_state["codigo_empresa_dom_val"] = codigo_empresa_dom

    st.divider()
    st.markdown(f'<p style="🤝 Convênios</p>', unsafe_allow_html=True)
    convenios = st.multifont-size:10px;font-weight:700;color:{TR_ORANGE};text-transform:uppercase;letter-spacing:1px;">📊 Parâmetrosselect("Entidades", label_visibility="collapsed",
        options=["SENAI SEFIP</p>', unsafe_allow_html=True)
    fap   = st.number_input("FAP",","SESI","SENAC","SESC","SEBRAE","SE           min_value=0.5, max_value=2.0, value=1.0, step=0.01NAR","SEST","SENAT","SESCOOP"])

    st.divider())
    delay = st.number_input("Intervalo (s)", min_value=0.3, max_value=5.0, value=1.
    st.markdown(f'<p style="font-size:10px;font-weight:700;color:{TR_ORANGE};text-transform:uppercase;letter-spacing:1px;">📅 Vig0, step=0.1)
    st.markdown(f'<p style="font-size:10px;font-weight:700;color:{TR_ORANGE};text-transform:uppercase;letter-spacing:1px;ência (FOVIGENCIAS)</p>', unsafe_allow_html=True)
    vigencia_data = st.date_input("Datamargin-top:12px;">🤝 Convênios</p>', unsafe_allow_ de início da vigência", value=date(2020,1,1))
    # ✅html=True)
    convenios = st.multiselect("Entidades", label_visibility="collapsed",
         Vigência no formato DD/MM/AAAA 00:00:00
    vigencia_str  = _options=["SENAI","SESI","SENACdata_para_vigencia(vigencia_data)
    descricao_vig = st.text_input("Descrição da vigência", value="Vigência","SESC","SEBRAE","SENAR","SEST","SENAT Inicial")

    st.divider()
    mapa  = st.session_state.get("","SESCOOP"])

    st.divider()
    st.markdown(f'<p style="font-size:10px;font-weight:700;color:{TR_MUNICIPIOS_MAP",{})
    debug = st.session_state.get("_mORANGE};text-transform:uppercase;letter-spacing:1px;">📅 Vigência (FOVun_debug",{})
    if mapa:
        st.success(f"✅ {len(mapa):IGENCIAS)</p>', unsafe_allow_html=True)
    vigencia_data = st.date_input("Data,} municípios carregados")
    else:
        st.warning(f"⚠️ Municípios não car de início da vigência", value=date(2020,1,1))
    vigencia_str  = vigencia_data.strregados\n{debug.get('erro_fatal','arquivo não encontrado')}")

    withftime("%Y-%m-%d") + " 00:00:00"
    descricao_vig = st.text_input("Descrição da vigência st.expander("🔍 Debug municípios"):
        st.json({"total":debug", value="Vigência Inicial")

    st.divider()
    mapa  = st.session.get("total",0),"col_codigo":debug.get("col_codigo",""),
                 "col_nome":debug.get("col_state.get("MUNICIPIOS_MAP",{})
    debug = st.session_state._nome",""),"col_uf":debug.get("col_uf",""),
                 "colunas":debug.get("cget("_mun_debug",{})
    if mapa:
        st.success(f"olunas",[]),"amostra":debug.get("amostra",[]),
                 "erros":debug.get("erros"✅ {len(mapa):,} municípios carregados")
    else:
        st.warning(f"⚠️ Munic,[])})
        teste_mun=st.text_input("Testar município",placeholder="SAípios não carregados\n{debug.get('erro_fatal','arquivo não encontrado')}")O PAULO")
        teste_uf=st.text_input("UF",placeholder="SP")
        if teste_mun and teste_uf:
            

    with st.expander("🔍 Debug municípios"):
        st.json({"totalcod_teste=buscar_codigo_municipio(teste_mun,teste_uf)
            if cod_teste: st.success":debug.get("total",0),"col_codigo":debug.get("col_codigo",""),
                 "col_(f"✅ Código: **{cod_teste}**")
            else:
                st.error("❌ Não encontnome":debug.get("col_nome",""),"col_uf":debug.get("col_uf",""),
                 "colurado")
                uf_n=_normalizar(teste_uf); nomenas":debug.get("colunas",[]),"amostra":debug.get("amostra",[]),
                 "er_n=_normalizar(teste_mun)
                sugestoes=[f"{n}→ros":debug.get("erros",[])})
        teste_mun=st.text_input("Testar município",{c}" for (u,n),c in list(mapa.items())placeholder="SAO PAULO")
        teste_uf=st.text_input("UF",placeholder="SP")
        if teste if u==uf_n and nome_n[:4] in n][:8_mun and teste_uf:
            cod_teste=buscar_codigo_municipio(teste_mun,teste_]
                if sugestoes: st.write("Sugestões:",sugestoes)

    if st.button("🔄uf)
            if cod_teste: st.success(f"✅ Código: **{cod_teste}**")
            else: Recarregar municípios"):
        for k in ["MUNICIPIOS_MAP","_mun_debug"]: st.session_state.
                st.error("❌ Não encontrado")
                uf_n=_normalizar(pop(k,None)
        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════teste_uf); nome_n=_normalizar(teste_mun)
                sugestoes=[f"{n════════════════════════════════════════════════════════════════
tab_lote, tab_individual, tab_tabela = st.tabs([
    "}→{c}" for (u,n),📋 Consulta em Lote", "🔍 Consulta Individual", "📖 Tabelac in list(mapa.items()) if u==uf_n and nome_n[:4] in n][: FPAS"
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — CONSULTA EM L8]
                if sugestoes: st.write("Sugestões:",sugestoes)

    ifOTE
# ══════════════════════════════════════════════════════════════════════════════
with tab_lote:
    st.markdown(f'<div class="tr-card st.button("🔄 Recarregar municípios"):
        for k in ["MUNICIPIOS_MAP","_mun_debug"]:"><div class="tr-card-title">📋 Passo 1 — Cole os CNPJs</div>', unsafe_allow_html=True) st.session_state.pop(k,None)
        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# TABS
    col_input,col_dica=st.columns([
# ══════════════════════════════════════════════════════════════════════════════
tab_lote, tab_individual, tab_tab3,1])
    with col_input:
        texto_cnpjs=st.text_area("CNPJs",ela = st.tabs([
    "📋 Consulta em Lote", "🔍 Consulta Individual", "📖label_visibility="collapsed",height=180,
            placeholder="Cole aqui os CNPJs — Tabela FPAS"
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — um por linha, vírgula, ponto-e-vírgula ou espaço.") CONSULTA EM LOTE
# ══════════════════════════════════════════════════════════════════════════════
with tab_lote:
    st.markdown(f'<div
    with col_dica:
        st.markdown(f"""
        <div style="background:{TR_CARD2};border: class="tr-card"><div class="tr-card-title">📋 Passo 1 — Cole os1px solid {TR_BORDER};border-radius:8px;padding:14px;font-size:11px;color:{TR_TEXT_MUTED};line CNPJs</div>', unsafe_allow_html=True)
    col_input,col_d-height:1.8;">
            <b style="color:{TR_ORANGE};">Formatos aceitos:</b><br>
            ✅ ica=st.columns([3,1])
    with col_input:
        texto_cnpjs=<code>00.000.000/0000-00</code><br>
            ✅st.text_area("CNPJs",label_visibility="collapsed",height=180,
            placeholder="Cole <code>00000000000000</code><br>
            ✅ Vírgula, ponto-e-vírgula,<br>&nbsp;&nbsp;&nbsp;quebra de linha, tab<br>
            ✅ Duplic aqui os CNPJs — um por linha, vírgula, ponto-e-vírgula ou espaço.")
    with col_dica:
        st.markdown(f"""
        <div style="background:{TR_CARD2};border:atas removidas
        </div>""", unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if texto_cnpjs and texto_cnpjs.strip():
        lista_cnpjs=extrair_cnpjs_do_texto(texto_cnpjs)
        if lista_cnpjs:
            validos  =[c for c in lista_cnpjs if validar_cnpj(c)]
            invalidos=[c for c in lista_cnpjs if not validar_cnpj(c)]
            chips_html="".join(
                f'<span class="cnpj-chip {"" if validar_cnpj(c) else "invalido"}">{"✅" if validar_cnpj(c) else "❌"} {c}</span>'
                for c in lista_cnpjs)
            err_html=f'· <span style="color:{TR_ERROR};">{len(invalidos)} inválido(s)</span>' if invalidos else ""
            st.markdown(f"""
            <div class="cnpj-preview-box">
                <div style="font-size:11px;color:{TR_TEXT_MUTED};text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px;">
                    {len(lista_cnpjs)} detect1px solid {TR_BORDER};border-radius:8px;padding:14px;font-size:11px;color:{TR_TEXT_MUTED};line-height:1.8;">
            <b style="color:{TR_ORANGE};">Formatos aceitos:</b><br>
            ✅ <code>00.000.000/0000-00</code><br>
            ✅ <code>00000000000000</code><br>
            ✅ Vírgula, ponto-e-vírgula,<br>&nbsp;&nbsp;&nbsp;quebra de linha, tab<br>
            ✅ Duplicatas removidas
        </div>""", unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if texto_cnpjs and texto_cnpjs.strip():
        lista_cnpjs=extrair_cnpjs_do_texto(texto_cnpjs)
        if lista_cnpjs:
            validos  =[c for c in lista_cnpjs if validar_cnpj(c)]
            invalidos=[c for c in lista_cnpjs if not validar_cnpj(c)]
            chips_html="".join(
                f'<span class="cnpj-chip {"" if validar_cnpj(cado(s) — <span style="color:{TR_SUCCESS};">{len(validos)} válido(s)</span>) else "invalido"}">{"✅" if validar_cnpj(c) else "❌"} {c}</span>'
                 {err_html}
                </div>{chips_html}
            </div>""", unsafe_allow_html=True)

            st.markdown("<brfor c in lista_cnpjs)
            err_html=f'· <span style="color:{TR_>", unsafe_allow_html=True)
            col_b1,col_b2=st.columns([3,1])
            with col_b1:
                inicERROR};">{len(invalidos)} inválido(s)</span>' if invalidos else ""
            st.markdown(f"""
            <div classiar=st.button(f"🔍 Passo 2 — Consultar e="cnpj-preview-box">
                <div style="font-size:11px;color:{TR_TEXT_MUTED};text Classificar {len(validos)} CNPJ(s)",
                    type="primary",use_container_width=True,-transform:uppercase;letter-spacing:.5px;margin-bottom:8px;">
                    {len(lista_cnpjs)} detectado(s) —disabled=len(validos)==0)
            with col_b2:
                if st.button("🗑️ Limpar",use_container_width=True <span style="color:{TR_SUCCESS};">{len(validos)} válido(s)</span> {err_html}
                ):
                    for k in ["resultados_proc","dados_brutos","seq</div>{chips_html}
            </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html_confirmada",
                              "seq_inicio_val","_txt=True)
            col_b1,col_b2=st.columns([3,1])
            with col_b1:
                inic","_xlsx","_vig"]:
                        st.session_state.pop(k,None)iar=st.button(f"🔍 Passo 2 — Consultar e
                    st.rerun()

            if iniciar and validos:
                resultados_proc=[ Classificar {len(validos)} CNPJ(s)",
                    type="primary",use_container_width=True,]; dados_brutos={}
                total=len(validos)
                progress=stdisabled=len(validos)==0)
            with col_b2:
                if st.button("🗑️ Limpar",use_container_width=True):
                    for.progress(0,text="Iniciando...")
                col_log,col_stat=st.columns([2,1])
                log k in ["resultados_proc","dados_brutos","seq_confirm_area=col_log.empty(); stat_area=col_stat.empty()
                logs,ok_nada",
                              "seq_inicio_val","_txt",",err_n=[],0,0

                for i,cnpj_raw_xlsx","_vig"]:
                        st.session_state.pop(k,None)
                    st.rerun()

            if inic in enumerate(validos):
                    progress.progress(int((i+1)/total*100),text=f"⏳ {iar and validos:
                resultados_proc=[]; dados_brutos={}
                totali+1}/{total} — {cnpj_raw}")
                    dados_rf=consultar_cnpj(cnpj_raw,delay=len(validos)
                progress=st.progress(0,text="Iniciando...")
                col=delay)

                    if dados_rf.get("erro"):
                        err_n+=1; logs.append(f"❌ {cnpj_raw_log,col_stat=st.columns([2,1])
                log_area=col_log.empty();}: {dados_rf['erro']}")
                        resultados_proc.append({
                            "cnpj":limpar_cnpj(cnpj_raw),"raz stat_area=col_stat.empty()
                logs,ok_n,err_n=[],ao_social":"","municipio":"","uf":"",
                            "cod_municipio_dom0,0

                for i,cnpj_raw in enumerate(validos):
                    progress.progress(int":"","cnae_codigo":"","fpas_descricao":"","codigo_fpas":"",
                            "((i+1)/total*100),text=f"⏳ {i+1}/{total} — {cnpj_raw}")
                    dados_rfcodigo_terceiro":"","perc_acid_trabalho":"","codigo_gps=consultar_cnpj(cnpj_raw,delay=delay)

                    if dados_rf.get("erro"):
                        ":"","codigo_gfip":"",
                            "_status":"ERRO_RF","_obs":dados_rf["erro"],
                        })
                        dados_brutos[i]=None
                    err_n+=1; logs.append(f"❌ {cnpj_raw}: {dados_rf['erro']}")
                        resultados_proc.append({
                            "cnpj":limpar_cnpj(cnpj_raw),"razao_social":"","municipio":"","uf":"",else:
                        simples=dados_rf.get("simples",False)
                        classif=classificar(dados_rf.get("cn
                            "cod_municipio_dom":"","cnae_codigo":"","fpas_descae_codigo",""),simples=simples,fap=fap,convenios=convenios)
                        status="ERRO_FPAS" if classif.get("erro") else "OK"
                        if statusricao":"","codigo_fpas":"",
                            "codigo_terceiro":"","perc_acid_trabalho":"","codigo_gps":"","codigo_gfip":"",
                            "_=="OK":
                            ok_n+=1; cod3=classif["codigo_terceiro"]
                            logsstatus":"ERRO_RF","_obs":dados_rf["erro"],
                        .append(f"✅ {cnpj_raw} | {dados_rf.get('razao_social','')[:26})
                        dados_brutos[i]=None
                    else:
                        simples=dados_rf.get("simples",False)]} | FPAS {classif['fpas']} | 3
                        classif=classificar(dados_rf.get("cnae_codigo",""),simples=simples,ºs {f'{cod3:04d}' if isinstance(cod3,int) else '-'}")
                        else:fap=fap,convenios=convenios)
                        status="ERRO
                            err_n+=1; logs.append(f"⚠️ {cnpj_raw} | {dados_rf.get('razao_social','')[:28]} | {_FPAS" if classif.get("erro") else "OK"
                        if status=="OK":
                            ok_n+=1;classif['erro']}")

                        municipio=dados_rf.get("municipio",""); uf=dados_rf.get("uf","") cod3=classif["codigo_terceiro"]
                            logs.append(f"✅ {cnpj_raw}
                        cod_mun=buscar_codigo_municipio(municipio,uf)
                        r_merged={ | {dados_rf.get('razao_social','')[:26]} | FPAS {
                            "cnpj":limpar_cnpj(cnpj_raw),
                            "razao_social":dados_rf.get("razao_social",""),
                            "cnclassif['fpas']} | 3ºs {fae_codigo":dados_rf.get("cnae_codigo",""),
                            "logradouro":dados_rf.get("logradouro",""),
                            "numero":dados_rf.get("'{cod3:04d}' if isinstance(cod3,int) else '-'}")
                        else:
                            err_n+=1; logs.append(f"numero",""),
                            "bairro":dados_rf.get("bairro",""),
                            "municipio":municipio,"uf":uf,
                            "cep":dados_rf.get⚠️ {cnpj_raw} | {dados_rf.get('razao_social','')[:28]} | {classif['erro']}")

                        municip("cep",""),
                            "data_inicio":dados_rf.get("data_inicio",""),
                            "codigo_fpas":classif.io=dados_rf.get("municipio",""); uf=dados_rf.get("uf","")
                        cod_mun=buscar_codigo_municipio(municipio,uf)get("fpas",""),
                            "codigo_terceiro":classif.get("codigo_terceiro",""),
                            "perc_acid_trabalho
                        r_merged={
                            "cnpj":limpar_cnpj(cnpj_raw),":classif.get("perc_acid_trabalho",""),
                            "codigo_gps":classif.get("codigo_gps",""),
                            "codigo_gfip":classif.get("codigo_gfip","
                            "razao_social":dados_rf.get("razao_social",""),
                            "cnae_codigo":dados_rf.get("cnae_codigo",""),
                            "log"),
                            "entidades":classif.get("entidades",[]),
                        }
                        dados_brutos[i]=r_merged
                        resultadosradouro":dados_rf.get("logradouro",""),
                            "numero":dados_rf.get("numero",""),
                            "bairro":dados_rf.get("bairro",""),
                            "municip_proc.append({
                            "cnpj":limpar_cnpj(cnpj_raw),
                            "razao_social":dados_rf.get("razao_social",""),
                            "municipio":io":municipio,"uf":uf,
                            "cep":dados_rf.get("cep",""),
                            "data_inicio":dados_rf.get("data_inicio",""),municipio,"uf":uf,
                            "cod_municipio_dom":cod_mun or "⚠️ não encontrado",
                            "cnae_codigo":dados_
                            "codigo_fpas":classif.get("fpas",""),
                            "codigo_tercrf.get("cnae_codigo",""),
                            "fpas_descricao":classif.get("fpas_descricao",""),
                            "codigo_fpas":classif.get("fpas","eiro":classif.get("codigo_terceiro",""),
                            "perc_acid_trabalho":classif.get("perc_acid_trabalho",""),
                            "codigo_gps":classif.get"),
                            "codigo_terceiro":(
                                f"{classif['codigo_terceiro']:04d}"
                                if isinstance(classif.("codigo_gps",""),
                            "codigo_gfip":classif.get("codigo_gfip",""),
                            "entidades":classif.get("entidades",[]),
                        }get("codigo_terceiro"),int)
                                else classif.get("codigo_terceiro","")
                            ),
                            "perc_acid_trabalho":classif.get("perc_acid_trabalho",""),
                
                        dados_brutos[i]=r_merged
                        resultados_proc.append({
                            "cnpj":limpar_cnpj(cnpj_raw),
                            "razao_            "codigo_gps":classif.get("codigo_gps",""),
                            "codigo_gfip":classif.get("codigo_gfip",""),
                            "_status":status,social":dados_rf.get("razao_social",""),
                            "municipio":municipio,"uf":uf,
                            "cod_municipio_dom":cod_mun or "
                            "_obs":classif.get("observacao","") if status=="OK" else classif.get("erro",""),
                        })⚠️ não encontrado",
                            "cnae_codigo":dados_rf.get("cnae_codigo",""),
                            "fpas_descricao":classif.get("fpas_

                    log_area.text_area("log","\n".join(logs[-descricao",""),
                            "codigo_fpas":classif.get("fpas",""),
                            "codigo_terceiro":(
                                f"{classif['12:]),height=240,label_visibility="collapsed")
                    stat_area.markdown(f"""codigo_terceiro']:04d}"
                                if isinstance(classif.get("codigo_terceiro"),int)
                                else classif.get("codigo_terceiro","")
                            ),
                    <div class="tr-metrics" style="flex-direction:column;">
                        {metric
                            "perc_acid_trabalho":classif.get("perc_acid_trabalho",""),
                            "codigo_gps":classif.get("codigo_gps",""),
                            "codigo_gfip":_card(i+1,"Processados")}
                        {metric_card(ok_n,"classif.get("codigo_gfip",""),
                            "_status":status,
                            "_obs":classif.get("observacao","") ifSucesso","success")}
                        {metric_card(err_n,"Erros","error")}
                    </div>""", unsafe status=="OK" else classif.get("erro",""),
                        })

                    log_area.text_area("_allow_html=True)

                progress.progress(100,text="✅ Concluído!")
                st.session_state["resultados_proc"]=resultados_proc
                st.log","\n".join(logs[-12:]),height=240,session_state["dados_brutos"]=dados_brutos
                for k in ["seq_confirmada","seq_inicio_val","_txt","_xlsx","_vig"]:
                    st.session_label_visibility="collapsed")
                    stat_area.markdown(f"""
                    <div class="tr-metrics" style="flexstate.pop(k,None)
        else:
            st.warning("⚠️ Nenhum CNPJ detectado. Verifique o-direction:column;">
                        {metric_card(i+1,"Processados formato.")

    if "resultados_proc" in st.session_state:
        resultados_proc=st.session_state["resultados_proc"]
        dados_brutos=st.session_state")}
                        {metric_card(ok_n,"Sucesso","success")}
                        {metric.get("dados_brutos",{})
        df_proc=pd.DataFrame(resultados_proc)
        ok_n=len(df_proc[df_proc["__card(err_n,"Erros","error")}
                    </div>""", unsafe_allow_html=True)

                progress.progress(100,text="✅ Concluído!")status"]=="OK"])
        err_n=len(df_proc[df_proc["_status"]!="OK"])

        st.markdown("---")
        st.markdown(
                st.session_state["resultados_proc"]=resultados_proc
                st.session_state["dados_brutos"]=dados_brutos
                for k in ["seqf"""
        <div class="tr-metrics">
            {metric_card(len(resultados_proc),"Total_confirmada","seq_inicio_val","_txt","_xlsx","_vig"]:
                    st.session_state.pop(k,None)
        else:
            st.warning")}
            {metric_card(ok_n,"Classificados","success")}
            {metric_card(err_n,"Erros","error")}
        </div>""",("⚠️ Nenhum CNPJ detectado. Verifique o formato.")

    if "resultados_proc" in st.session_state:
         unsafe_allow_html=True)
        cols_show=[c for c in df_proc.columns if not c.startswith("_")]resultados_proc=st.session_state["resultados_proc"]
        dados_brutos=st.session_state.get("dados_brutos",{})
        df_proc
        st.dataframe(df_proc[cols_show],use_container_width=True,hide_index=True,height=300)

        st=pd.DataFrame(resultados_proc)
        ok_n=len(df_proc[df_proc["_status"]=="OK"])
        err_n=len(df_proc[df_proc["_status"].markdown(f"""
        <div class="tr-card">
            <div class="tr-card-title">🏷️ Passo 3!="OK"])

        st.markdown("---")
        st.markdown(f"""
        <div class="tr-metrics"> — Código de Serviço, Tipo e Revisão</div>
            
            {metric_card(len(resultados_proc),"Total")}
            {metric_card(ok_n,"Classificados","<div style="font-size:12px;color:{TR_TEXT_MUTED};margin-bottom:16px;">success")}
            {metric_card(err_n,"Erros","error")}
        </div>""", unsafe_allow_html=True)
        cols_show
                Informe o número inicial da sequência e clique em =[c for c in df_proc.columns if not c.startswith("_")]
        st.dataframe(df_proc[cols_show],use_container_width=True,hide<b style="color:{TR_ORANGE};">Confirmar</b>.<br>
                <b_index=True,height=300)

        st.markdown(f"""
        <div class="tr-card">
            <div class=">Código de Serviço = Codigo_Servicos = Codigo_tr-card-title">🏷️ Passo 3 — Código de Serviço,eSocial</b>
            </div>
        """, unsafe_allow_html=True)

        col_seq1 Tipo e Revisão</div>
            <div style="font-size:12px;color:{TR_TEXT_MUTED};margin,col_seq2,_=st.columns([1,1,2])
        with col_seq1:
            seq_inicio-bottom:16px;">
                Informe o número inicial da sequência e cl=st.number_input("🔢 Número inicial da sequência",
                min_value=1,max_value=999999,ique em <b style="color:{TR_ORANGE};">Confirmar</b>.<br>
                
                value=st.session_state.get("seq_inicio_val",1),
                step=1,key="seq_inicio_input")
            confir<b>Código de Serviço = Codigo_Servicos = Codigomar_seq=st.button("✅ Confirmar sequência",use_container_width=True)
            if confirmar_seq:
                for_eSocial (foservicos)</b>
            </div>
         i in range(len(resultados_proc)): st.session_state.pop(f"cod_srv_{i}",None)""", unsafe_allow_html=True)

        col_seq1,col_seq2,_=
                st.session_state["seq_inicio_val"]=seq_inicio
                st.session_state["seq_confirmada"]=True
                st.rerun()
        with col_seq2:st.columns([1,1,2])
        with col_seq1:
            seq_inicio=st.number_input("🔢
            seq_val=st.session_state.get("seq_inicio_val",None)
            if seq_val and st Número inicial da sequência",
                min_value=1,max_value=999999,
                value=st.session_state.get("seq_inicio_val",1),
                step=1,.session_state.get("seq_confirmada"):
                st.markdown(f"""
                <div style="background:{TR_CARD2};border:1px solidkey="seq_inicio_input")
            confirmar_seq=st.button("✅ Confirmar sequência",use_container_width=True) {TR_BORDER};border-radius:8px;
                            padding:12px;margin-top:28px;font-size:11
            if confirmar_seq:
                for i in range(len(resultados_proc)): st.session_state.poppx;color:{TR_TEXT_MUTED};">
                    📌 Sequência confirmada:<br>
                    <b style="color:{TR_ORANGE(f"cod_srv_{i}",None)
                st.session_state["seq_inicio_val"]=seq_inicio
                st.session};font-size:14px;">{seq_val} → {seq_val+len(resultados_proc)-1}</b>
                </div_state["seq_confirmada"]=True
                st.rerun()
        with col_seq2:
            seq_val=st.session_state.get("seq_inicio>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        if_val",None)
            if seq_val and st.session_state.get("seq_confirmada"):
                st.markdown(f"""
                <div style not st.session_state.get("seq_confirmada"):
            st.info("ℹ️ Informe e confir="background:{TR_CARD2};border:1px solid {TR_BORDER};border-radius:8px;
                            padding:12px;margin-top:28me o número inicial da sequência para exibir os campos de Código de Serviço.")
        else:
            seqpx;font-size:11px;color:{TR_TEXT_MUTED};">
                    📌 Sequência confirmada:<br>
                    <b style_inicio=st.session_state["seq_inicio_val"]
            st.markdown(f"""
            <div style="display:grid;grid-template-columns:2fr 1fr 1fr;gap:8px;
                        padding:6px 12px;background:{TR_="color:{TR_ORANGE};font-size:14px;">{seq_val} → {seq_val+len(resultados_proc)-1}</b>
                </div>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_CARD};border-radius:6px;
                        font-size:10px;font-weight:700;color:{TR_html=True)

        if not st.session_state.get("seq_confirmada"):
            st.info("ℹ️ Informe eORANGE};
                        text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px;">
                <div confirme o número inicial da sequência para exibir os campos de Código de Serviço.")
        else:
            seq_inicio=st.session_state["seq_inicio_val"]
            st.markdown(f"""
            <div style="display:grid;grid-template-columns:2fr 1fr 1fr;gap:8px;
                        padding:6px 12px;background:{TR_CARD};border-radius:6px;
                        font-size:10px;font-weight:700;color:{TR_ORANGE};
                        text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px;">
                <div>Empresa</div><div>Cód. Serviço</div><div>Tipo</div>
            </div>""", unsafe_allow_html=True>Empresa</div><div>Cód. Serviço</div><div>Tipo</div>
            </div>""", unsafe_allow_html=True)

            tipos_selecionados={}; codigos_servico={}

            for idx,row in enumerate(resultados_proc):
                status=row.get("_status","")
                mun=row.get("municipio","") or ""; uf=row.get("uf","") or ""
                cod_mun=row.get("cod_municipio_dom","")
                cor_mun=TR_SUCCESS if cod_mun and "não")

            tipos_selecionados={}; codigos_serv not in str(cod_mun) else TR_ERROR
                cor_st=TR_SUCCESS if status=="OK" else TR_ERROR
                icico={}

            for idx,row in enumerate(resultados_proc):
                statusone="✅" if status=="OK" else "❌"
                seq_auto=seq_inicio+idx

                col=row.get("_status","")
                mun=row.get("municipio","") or ""; uf=row.get("_info,col_cod_srv,col_tipo=st.columns([2,1,1])
                with col_info:
                    st.markdown(f"""uf","") or ""
                cod_mun=row.get("cod_municipio_dom","")
                cor_m
                    <div class="tipo-row">
                        <div style="font-size:12px;font-weight:700;color:{TR_TEXT}un=TR_SUCCESS if cod_mun and "não" not in str;">
                            {icone} <code style="color:{TR_ORANGE};">{row['(cod_mun) else TR_ERROR
                cor_st=TR_SUCCESS if status=="OK" else TR_ERROR
                icone="cnpj']}</code>
                            &nbsp; {(row.✅" if status=="OK" else "❌"
                seq_auto=seq_inicio+idx

                col_info,col_codget('razao_social','') or '—')[:40]}
                        </div>
                        <div style="_srv,col_tipo=st.columns([2,1,1])
                with col_info:
                    st.markdown(f"""
                    <div classfont-size:10px;color:{TR_TEXT_MUTED};margin-top:3px;">
                            FPAS <b="tipo-row">
                        <div style="font-size:12px;font-weight:700;color:{TR_TEXT} style="color:{TR_ORANGE};">{row.get('codigo_fpas','—')}</b>
                            &nbsp;·;">
                            {icone} <code style="color:{TR_ORANGE};">{row['&nbsp; {mun}/{uf}
                            &nbsp;·&nbsp; Cód.Mcnpj']}</code>
                            &nbsp; {(row.un: <b style="color:{cor_mun};">{cod_mun or '⚠️'}</b>
                            &nbsp;·&nbsp; <spanget('razao_social','') or '—')[:40]}
                        </div>
                        <div style="font-size:10px;color:{TR_TEXT style="color:{cor_st};">{status}</span>
                        </div>
                    </div>""", unsafe_allow_html=True)
                with col_cod_srv:_MUTED};margin-top:3px;">
                            FPAS <b style="color:{TR_ORANGE
                    cod_srv=st.number_input(f"Cód. Serviço #{idx};">{row.get('codigo_fpas','—')}</b>
                            &nbsp;·&nbsp; {mun}/{uf}
                            +1}",
                        min_value=1,max_value=999999,value=seq_auto,step=1,
                        key=f"cod_srv_{idx}",label&nbsp;·&nbsp; Cód.Mun: <b style="color:{cor_mun};">{cod_mun or_visibility="collapsed")
                    codigos_servico[idx]=int(cod_srv)
                with col_tipo:
                    tipos_selecionados[idx '⚠️'}</b>
                            &nbsp;·&nbsp; <span style="color:{cor_st};">{status}</span>
                        ]=st.selectbox(f"Tipo #{idx+1}",
                        options=list(TIPOS_EMPRESA.keys()),
                        format_func=lambda k</div>
                    </div>""", unsafe_allow_html=True)
                with col_cod_srv:
                    cod_srv=st.number: f"{k} — {TIPOS_EMPRESA[k]}",
                        key=f"tipo_{idx}_input(f"Cód. Serviço #{idx+1}",
                        min_value=1,max_value=999999,value",label_visibility="collapsed")

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown(f'<div class="tr-card">=seq_auto,step=1,
                        key=f"cod_srv_{idx}",label_visibility="collapsed")
                    codigos_servico[idx]=<div class="tr-card-title">⬇️ Passo 4 — Gerar e Baixar Arquivos</div></div>', unsafe_allow_htmlint(cod_srv)
                with col_tipo:
                    tipos_selecionados[idx]=st.selectbox(f"Tipo #{idx+1}",
                =True)

            col_d1,_=st.columns(2)
            with col_d1:
                if st.button("        options=list(TIPOS_EMPRESA.keys()),
                        format_func=lambda k: f"{k}⚙️ Gerar Arquivos",type="primary",use_container_width=True):
                    codigo_empresa_ — {TIPOS_EMPRESA[k]}",
                        key=f"tipo_{idx}",label_visibility="collapsed")

            st.markdown("<br>", unsafe_allow_dom=st.session_state.get("codigo_empresa_dom_val",1)
                    linhas_fo=[]; linhtml=True)
            st.markdown(f'<div class="tr-card"><div class="tr-card-has_vig=[]

                    for idx,row in enumerate(resultados_proc):
                        cod_srv =title">⬇️ Passo 4 — Gerar e Baixar Arquivos</div>codigos_servico.get(idx,seq_inicio+idx)
                        tipo_cod=tipos_selecionados.get(idx,1</div>', unsafe_allow_html=True)

            col_d1,_=st.columns(2)
            with col_d1:
                if)
                        r_merged=dados_brutos.get(idx)

                        # foservicos st.button("⚙️ Gerar Arquivos",type="primary",use_container_width=True):
                    codigo_empresa_dom=st..txt
                        if r_merged:
                            linha_fo=montar_linha_dominiosession_state.get("codigo_empresa_dom_val",1)
                    linhas_fo=[]; linhas_(r_merged,tipo_cod=tipo_cod,
                                cod_servico=cod_srv,codigo_empresa=codigo_empresa_dom)vig=[]

                    for idx,row in enumerate(resultados_proc):
                        cod_srv =codigos_servico.
                        else:
                            linha_fo=[""]* 25
                            get(idx,seq_inicio+idx)
                        tipo_cod=tipos_selecionados.get(idx,1)
                        rlinha_fo[0]=codigo_empresa_dom; linha_fo[1]=cod_srv
                            linha_fo[2]=row_merged=dados_brutos.get(idx)

                        # foservicos.txt["cnpj"]; linha_fo[3]=1
                            linha_fo[17]=cod_srv; linha_fo[18]=1;
                        if r_merged:
                            linha_fo=montar_linha_dominio(r_merged,tipo_cod=tipo_cod,
                                 linha_fo[19]=tipo_cod
                            linha_fo[21]="01/01/2020cod_servico=cod_srv,codigo_empresa=codigo_empresa_dom)
                        else:
                            linha_fo=[""; linha_fo[22]=1; linha_fo[23]=cod_srv
                        linhas_fo.append(linha_fo)

                        # fov"]* 25
                            linha_fo[0]=codigoigencias_servicos.txt
                        if r_merged:
                            cod_mun__empresa_dom; linha_fo[1]=cod_srv
                            linha_fo[2]=row["cnpj"]; linha_fo[3]=1v=buscar_codigo_municipio(
                                r_merged.get("municipio",""),r_merged.get("uf",""))
                            linha
                            linha_fo[17]=cod_srv; linha_fo[18]=1; linha_fo[19]=tipo_cod
                            linha_fo[21_vig=montar_linha_vigencia(
                                r_merged,cod_servico=cod_srv,tipo_cod=tipo_cod,
                                codigo_empresa=]="2020-01-01"; linha_fo[22]=1; linha_fo[23]=cod_srv
                        codigo_empresa_dom,
                                vigencia=vigencia_str,
                                descricao_viglinhas_fo.append(linha_fo)

                        # fovigencias_servicos.txt
                        if r_merged:
                            encia=descricao_vig,
                                cod_mun=cod_mun_v)
                        else:
                            linha_vig=_linhacod_mun_v=buscar_codigo_municipio(
                                r_merged.get("municipio",""),r_vigencia_vazia(
                                row["cnpj"],cod_srv,tipo_cod,codigo_empresa_dom,
                                _merged.get("uf",""))
                            linha_vig=montar_linha_vigencia(
                                r_merged,cod_servvigencia_str,descricao_vig)
                        linhas_vig.append(linha_vig)

                    df_conf=pd.DataFrame(linhasico=cod_srv,tipo_cod=tipo_cod,
                                codigo_empresa=codigo_empresa_dom,
                                vigencia=vigencia_str_fo,columns=COLUNAS_LEIAUTE)
                    df_conf[",descricao_vigencia=descricao_vig,
                                cod_mun=cod_mun_v)
                        else:
                            linha_vig=_status"]=[r["_status"] for r in resultados_proc]
                    df_err=df_linha_vigencia_vazia(
                                row["cnpj"],cod_srv_conf[df_conf["_status"]!="OK"]
                    cols_xl=[c for c in df_conf.columns if c!="_,tipo_cod,codigo_empresa_dom,
                                vigencia_str,descricao_vig)
                        linhas_vig.append(linha_vig)

                    dfstatus"]

                    st.session_state["_txt"] =gerar_txt_le_conf=pd.DataFrame(linhas_fo,columns=COLUNAS_LEIAUTEiaute(linhas_fo)
                    st.session_state["_xlsx"]=gerar_excel_conferencia()
                    df_conf["_status"]=[r["_status"] for r in result
                        df_conf[cols_xl+["_status"]],
                        df_err[cols_xl+["_status"]] if lenados_proc]
                    df_err=df_conf[df_conf["_status"]!="OK"]
                    cols_xl=[(df_err)>0 else None)
                    st.session_state["_vig"] =gerar_txt_vigencias(linhas_vig)
                    st.successc for c in df_conf.columns if c!="_status"]

                    st.session_state["_("✅ Arquivos gerados com sucesso!")

            if "_txt" in st.session_state andtxt"] =gerar_txt_leiaute(linhas_fo)
                    st.session_ "_vig" in st.session_state:
                col_dl1,col_dl2,col_dl3=st.columnsstate["_xlsx"]=gerar_excel_conferencia(
                        df_conf[cols_xl+["_status"]],
                (3)
                with col_dl1:
                    st.download_button(
                        "📄 foservicos.txt\        df_err[cols_xl+["_status"]] if len(df_err)>0 else None)
                    st.session_n(Importação Domínio · 25 col · TAstate["_vig"] =gerar_txt_vigencias(linhas_vig)
                    st.success("✅ Arquivos gerados com sucesso!")B)",
                        data=st.session_state["_txt"],
                        file_name="foservicos.txt",
                        mime="text/plain",
                        use_container_width=True,

            # Downloads — protegidos com
                    )
                with col_dl2:
                    st.download_button(
                        "📄 fovigencias_servicos.txt\n(Vigências verificação conjunta
            if "_txt" in st.session_state and "_vig" in st.session_ · TAB)",
                        data=st.session_state["_vig"],
                        file_name="fovigencias_servicos.txt",
                        mime="text/plain",
                        use_containerstate:
                col_dl1,col_dl2,col_dl3=st.columns(3)
                with col_dl1:
                    st_width=True,
                    )
                with col_dl3:
                    st.download_button(
                        "📊 Excel Conferência\n(.xlsx format.download_button(
                        "📄 foservicos.txt\n(Importação Domínio ·ado)",
                        data=st.session_state["_xlsx"],
                        file_name="dominio_conferencia.xlsx",
                        mime="application/vnd.openxmlform 25 col · TAB)",
                        data=st.session_state["_txt"],
                        file_name="foservicos.ats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

# ══════════════════════════════════════════════════════════════════════════════txt",
                        mime="text/plain",
                        use_container_width=True,
                    )
                with col_dl2:
                    st.download_button(
                        "📄 fovig
# TAB 2 — CONSULTA INDIVIDUAL
# ══════════════════════════════════════════════════════════════════════════════
with tab_individual:
    st.markdown('encias_servicos.txt\n(Vigências · TAB)",
                        data=st.session_state["_vig"],
                        file<div class="tr-card"><div class="tr-card-title">🔍 Consulta Individual de CNPJ</div>', unsafe_allow_html=True)_name="fovigencias_servicos.txt",
                        mime="text/plain",
                        use_container_width=True,
                    )
                with col_dl3:
                    st.download_button(
                
    col_inp,col_btn=st.columns([3,1])
    with col_inp:
        cnpj_input=st.text_input("CNPJ",placeholder        "📊 Excel Conferência\n(.xlsx formatado)",
                        data=st.session_state["_xlsx"],
                        file_name="dom="00.000.000/0000-00",label_visibility="collapsed")
    with col_btn:
        buscar=st.button("🔍inio_conferencia.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

# Consultar",type="primary",use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if buscar and ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — CONSULTA INDIVIDUAL
# ══════════════ cnpj_input:
        with st.spinner("🔄 Consultando..."):
            dados_rf=consultar_cn════════════════════════════════════════════════════════════════
with tab_individual:
    st.markdown('<div class="tr-card"><div class="tr-card-title">🔍 Consulta Individualpj(cnpj_input,delay=0.3)
        if dados_rf.get("erro"):
            st.error de CNPJ</div>', unsafe_allow_html=True)
    col_inp,col_btn=st.columns([3,1])
    with col(f"❌ {dados_rf['erro']}")
        else:
            simples=dados_rf.get("simples",False)
            classif=classificar(_inp:
        cnpj_input=st.text_input("CNPJ",placeholder="00.000.000/0000-00",label_visibility="collapsed")
    dados_rf.get("cnae_codigo",""),simples=simples,fap=fap,convenios=convenios)
            mun=dados_rf.get("municipio","");with col_btn:
        buscar=st.button("🔍 Consultar",type="primary",use_container_width=True)
    st uf_val=dados_rf.get("uf","")
            cod_mun=buscar_codigo_municipio(mun,uf_val)
            st.markdown('</div>', unsafe_allow_html=True)

    if buscar and cnpj_input:
        with st.spinner(".markdown(f"""
            <div class="tr-card"><div class="tr-card-title">🏢 Dados</div>
            <div class🔄 Consultando..."):
            dados_rf=consultar_cnpj(cnpj_input,delay=0.="result-grid">
                {result_item("CNPJ",dados3)
        if dados_rf.get("erro"):
            st.error(f"❌ {dados_rf['_rf["cnpj"])}
                {result_item("Razão Social",dados_rf["razao_social"],True)}
                {result_item("Situação",dados_rf.get("situacao","—"))}
                {result_item("Simples Nacional","✅ Serro']}")
        else:
            simples=dados_rf.get("simples",False)
            classif=classificar(dados_rf.get("cnae_codigo",""),simples=simples,fap=fap,convenios=convenios)
            mun=dados_rf.get("municipio",""); uf_val=dados_rf.get("IM" if simples else "❌ NÃO")}
                {result_item("Município/UF",f"{mun}/{uf","")
            cod_mun=buscar_codigo_municipio(mun,uf_val)
            st.markdown(f"""
            <div class="tr-card"><div class="truf_val}")}
                {result_item("Cód. Município (Domínio)",cod_mun or "-card-title">🏢 Dados</div>
            <div class="result-grid">
                {result⚠️ não encontrado",True)}
                {result_item("CNAE",dados_rf.get("cnae_item("CNPJ",dados_rf["cnpj"])}
                {result_item("Razão Social",dados_rf["razao_social_codigo","—"),True)}
            </div></div>""", unsafe_allow_html=True)"],True)}
                {result_item("Situação",dados_rf.get("situacao","—"))}
                {result_
            if not classif.get("erro"):
                cod3=classif["codigo_terceiro"]
                item("Simples Nacional","✅ SIM" if simples else "❌ NÃO")}
                {result_item("st.markdown(f"""
                <div class="tr-card"><div class="tr-card-title">📋 Classificação FPAS</div>
                <div class="result-grid">Município/UF",f"{mun}/{uf_val}")}
                {result_item("Cód. Município
                    {result_item("FPAS",classif["fpas"],True)}
                    {result_item("Descrição",classif["fp (Domínio)",cod_mun or "⚠️ não encontrado",True)}
                {result_item("as_descricao"])}
                    {result_item("Cód. Terceiros",f"{cod3:04d}" if isinstance(cod3,int) else "CNAE",dados_rf.get("cnae_codigo","—"),True)}
            </div></div>""—",True)}
                    {result_item("RAT",f"{classif['perc_acid_trabalho']}%")}
                    {result_item("GPS", unsafe_allow_html=True)
            if not classif.get("erro"):
                cod",classif["codigo_gps"])}
                    {result_item("GFIP",classif["codigo_gfip"])}
                </div>3=classif["codigo_terceiro"]
                st.markdown(f"""
                <div class="tr-card"><div class="tr-card-title">📋 Classific</div>""", unsafe_allow_html=True)
            else:
                st.warning(f"⚠️ {classif['erro']}")

#ação FPAS</div>
                <div class="result-grid">
                    {result_item("FPAS",classif["fpas ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — TABELA FPAS
# ══════════════════════════════════════"],True)}
                    {result_item("Descrição",classif["fpas_descricao"])}
                    {result_item("Cód. Terceiros════════════════════════════════════════
with tab_tabela:
    st.markdown('<div class="tr-card"><div class="tr-card-title">📖 Tabela FPAS",f"{cod3:04d}" if isinstance(cod3,int) else "—",True)}
                    {result_item("RAT / Terceiros</div>', unsafe_allow_html=True)
    ref_rows=[]
    for fp",f"{classif['perc_acid_trabalho']}%")}
                    {result_item("GPS",classif["codigo_gps"])}
                    {result_as,(terc,rat,gps,gfip,desc) in FPAS_CONFIGitem("GFIP",classif["codigo_gfip"])}
                </div></div>""", unsafe_allow_html=True)
            else:
                st.warning.items():
        ents=decodificar_terceiros(terc); total_pct=sum(e["aliquota"] for e in ents)
        sig(f"⚠️ {classif['erro']}")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — TABELA FPlas=" + ".join(e["sigla"] for e in ents) if ents else "—"
        ref_rows.append({"AS
# ══════════════════════════════════════════════════════════════════════════════
with tab_tabela:
    st.markdown('<div class="tr-card"><div class="tr-card-title">📖 Tabela FPAS / Terceiros</div>', unsafe_allow_html=True)
    ref_rowsFPAS":fpas,"Descrição":desc,"Cód. Terceiros":f"{terc:04d}",
            "Ent=[]
    for fpas,(terc,rat,gps,gfip,descidades":siglas,"Total 3ºs (%)":f"{total_pct:.1f}%",
            "RAT Base () in FPAS_CONFIG.items():
        ents=decodificar_terc%)":f"{rat}%","Cód. GPS":gps,"Cód. GFIP":gfip})
    df_ref=pd.DataFrame(ref_rows)
    busca_fpas=st.text_input("🔎 Filtrar",placeholder="Ex: 833 ou Construção")
    if busca_fpas:
        mask=(df_ref["FPAS"].astype(str).str.contains(busca_fpas,case=False)
              |df_ref["Descrição"].str.contains(busca_fpas,case=False)
              |df_ref["Entidades"].str.contains(busca_fpas,case=False))
        df_ref=df_ref[mask]
    st.dataframe(df_ref,use_container_width=True,hide_index=True,height=480)eiros(terc); total_pct=sum(e["aliquota"] for e in ents)
        siglas=" + ".join(e["sigla"] for e in ents) if ents else "—"
        ref_rows.append({"FPAS":fpas,"Descrição":desc,"Cód. Terceiros":f"{terc:04d}",
            "Entidades":siglas,"Total 3ºs (%)":f"{total_pct:.1f}%",
            "RAT Base (%)":f"{rat}%","Cód. GPS":gps,"Cód. GFIP":gfip})
    df_ref=pd.DataFrame(ref_rows)
    busca_fpas=st.text_input
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown(f"""
<div class="tr-footer">
    <span("🔎 Filtrar",placeholder="Ex: 833 >DOMÍNIO SISTEMAS</span> &nbsp;·&nbsp; Thomson Reuters &ou Construção")
    if busca_fpas:
        mask=(df_ref["FPAS"].asnbsp;·&nbsp;
    Classificador FPAS / Terceiros / SEFIPtype(str).str.contains(busca_fpas,case=False)
              |df_ref["Descrição"].str.contains(busca_fp &nbsp;·&nbsp; IN RFB nº 971/2009as,case=False)
              |df_ref["Entidades"].str.contains(busca_fpas,case=False))
        df_ref=df_ref[mask]
     &nbsp;·&nbsp; <span>v8.2</span>
</div>st.dataframe(df_ref,use_container_width=True,hide_index=True,height=480)
    st.markdown('</div>', unsafe_allow_html=""", unsafe_allow_html=True)
